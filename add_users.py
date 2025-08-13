import logging
from db import init_db, add_user
import sqlite3

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

init_db()

users = [
    {"username": "12345678", "password": "password123"},
    {"username": "87654321", "password": "pass456"},
    {"username": "344", "password": "mypassword"},
    # add more users here
]

def export_users_to_excel(users, filename='user_logins_backup.xlsx'):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from werkzeug.security import generate_password_hash
    try:
        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "User Logins"
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers with styling
        headers = ['Username', 'Password Hash']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row_num, user in enumerate(users, 2):
            ws.cell(row=row_num, column=1, value=user['username'])
            ws.cell(row=row_num, column=2, value=generate_password_hash(user['password'], method='pbkdf2:sha256'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        wb.save(filename)
        logging.info(f'User login info exported to {filename}')
    except Exception as e:
        logging.error(f'Failed to export users to Excel: {e}')

added_count = 0
exists_count = 0

for user in users:
    try:
        if add_user(user["username"], user["password"]):
            logging.info(f'User {user["username"]} added.')
            added_count += 1
        else:
            logging.warning(f'User {user["username"]} already exists.')
            exists_count += 1
    except Exception as e:
        logging.error(f'Error adding user {user["username"]}: {e}')

logging.info(f'Total users processed: {len(users)}')
logging.info(f'Users added: {added_count}')
logging.info(f'Users already existed: {exists_count}')

export_users_to_excel(users)

# Debug script: Print all forms in the exception_forms table
with sqlite3.connect('forms.db', timeout=10) as conn:
    c = conn.cursor()
    c.execute('SELECT * FROM exception_forms')
    rows = c.fetchall()
    columns = [desc[0] for desc in c.description]
    print(f"Found {len(rows)} forms in exception_forms table:")
    for row in rows:
        form = dict(zip(columns, row))
        print(form)