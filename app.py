import os
from dotenv import load_dotenv
from flask import Flask, request, jsonify, g, Response
from flask_cors import CORS
from model import train_model
from db import init_db, add_user, check_user
import re
from db import init_exception_form_db, store_exception_form
from exception_codes import exception_codes
import sqlite3
import csv
import requests
import google.generativeai as genai
import datetime
import pdfplumber
from PIL import Image
import io
import json
from typing import Dict, Any, List, Tuple

# Load environment variables from .env file
load_dotenv()

# Configuration flags for extraction optimization
PURE_GEMINI_EXTRACTION = False  # Set to True to use pure extraction
ENHANCED_FORM_DETECTION = True  # Enable advanced form detection for maximum overtime slip extraction
MAX_SEGMENTS_PER_PAGE = 6  # Maximum number of segments to extract per PDF page

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

model = train_model()
init_db()
init_exception_form_db()    
# Enable WAL mode for better SQLite concurrency
with sqlite3.connect('forms.db', timeout=10) as conn:
    conn.execute('PRAGMA journal_mode=WAL;')

# Utility: Google Gemini extraction
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)
    gemini_model = genai.GenerativeModel('gemini-1.5-flash')
else:
    gemini_model = None

def _segment_similarity(segment1_data, segment2_data):
    """
    Simple similarity check to avoid processing duplicate segments.
    Returns True if segments are similar enough to be considered duplicates.
    """
    try:
        # Simple byte comparison for exact duplicates
        if segment1_data == segment2_data:
            return True
        
        # For more sophisticated similarity, we could implement image hashing
        # For now, just check if they're the same size and have similar byte patterns
        if len(segment1_data) == len(segment2_data):
            # Check if they're very similar (90%+ same bytes)
            same_bytes = sum(1 for a, b in zip(segment1_data, segment2_data) if a == b)
            similarity = same_bytes / len(segment1_data)
            return similarity > 0.9
        
        return False
    except Exception:
        return False

def detect_multiple_forms_in_document(file_path, form_type):
    """
    Advanced detection of multiple forms in a single document.
    Returns a list of detected form regions.
    """
    try:
        if file_path.lower().endswith('.pdf'):
            # For PDFs, use the existing segmentation logic
            return None  # Let the main logic handle it
        
        # For images, try to detect multiple form regions
        from PIL import Image, ImageEnhance
        import numpy as np
        
        img = Image.open(file_path)
        width, height = img.size
        
        # Convert to grayscale for analysis
        gray_img = img.convert('L')
        
        # Look for horizontal lines that might separate forms
        # This is a simple heuristic - look for rows with low pixel values (dark lines)
        form_regions = []
        
        # Start with the full image
        form_regions.append((0, height))
        
        # Look for potential horizontal separators
        for y in range(height // 4, height * 3 // 4, height // 8):
            # Check if this row has a dark line (potential separator)
            row_pixels = [gray_img.getpixel((x, y)) for x in range(0, width, width // 10)]
            avg_brightness = sum(row_pixels) / len(row_pixels)
            
            # If this row is significantly darker than surrounding rows, it might be a separator
            if avg_brightness < 100:  # Threshold for dark lines
                # Check if we can split here
                if y > height // 4 and y < height * 3 // 4:
                    # Split the image at this point
                    form_regions = [
                        (0, y),
                        (y, height)
                    ]
                    break
        
        # If we found potential splits, create segments
        if len(form_regions) > 1:
            segments = []
            for start_y, end_y in form_regions:
                if end_y - start_y > height // 3:  # Only if segment is large enough
                    segment = img.crop((0, start_y, width, end_y))
                    segments.append(segment)
            return segments
        
        return None
        
    except Exception as e:
        print(f"Error detecting multiple forms: {e}")
        return None

def gemini_extract_file_details(file_path, prompt=None, form_type=None):
    """
    Uses Google Gemini to extract details from a file (PDF or image).
    Optimized for maximum overtime slip extraction.
    Returns Gemini's response text or None if not configured.
    """
    if not gemini_model:
        print("Gemini API key not set. Skipping Gemini extraction.")
        return None
    
    # Enhanced prompt for maximum overtime slip extraction
    if prompt is None:
        if form_type == 'hourly':
            prompt = """You are an expert at extracting overtime slip information from NYC Transit forms. 
            
            CRITICAL: Look for MULTIPLE overtime entries on this form. Each form may contain several overtime slips.
            
            Extract ALL overtime information you can find, including:
            - Employee details (pass number, name, title, RDOS)
            - Multiple overtime entries with times, locations, and reasons
            - Exception codes and descriptions
            - Line locations and run numbers
            - All time fields (exception time from/to, overtime, bonus, night differential)
            - Job numbers and account information
            
            If you see multiple overtime entries, structure them as separate entries in an 'entries' array.
            Be thorough - extract every piece of overtime information visible on the form.
            
            Return as structured JSON with all fields populated."""
        elif form_type == 'supervisor':
            prompt = """You are an expert at extracting supervisor overtime authorization information from NYC Transit forms.
            
            CRITICAL: This image may contain PARTIAL forms, COMPLETE forms, or MULTIPLE forms. Extract EVERYTHING you can see.
            
            AGGRESSIVE EXTRACTION RULES:
            1. Look for ANY text, numbers, or form elements - even if incomplete
            2. Extract employee details from ANY visible section (pass number, name, title, RDOS)
            3. Look for overtime information in ANY format (times, hours, locations)
            4. Find job numbers, RC numbers, account numbers wherever they appear
            5. Extract ANY reason checkboxes or text that might indicate overtime reasons
            6. Look for supervisor/superintendent signatures, dates, or pass numbers
            7. Capture ANY comments or additional information
            
            EXTRACTION STRATEGY:
            - If you see a complete form: extract all fields
            - If you see a partial form: extract whatever fields are visible
            - If you see multiple forms: structure them in an 'entries' array
            - If you see form fragments: extract what you can and mark incomplete fields as null
            - If you see ANY text that might be form-related: include it
            
            BE EXTREMELY THOROUGH - extract every single piece of information visible, even if it seems incomplete or unclear.
            
            Return as structured JSON with all visible fields populated. Use null for missing fields, but extract everything you can see."""
        else:
            prompt = """You are an expert at extracting overtime slip information from NYC Transit forms.
            
            CRITICAL: Look for MULTIPLE overtime entries on this form. Each form may contain several overtime slips.
            
            Extract ALL overtime information you can find, including:
            - Employee details (pass number, name, title, RDOS)
            - Multiple overtime entries with times, locations, and reasons
            - Exception codes and descriptions
            - Line locations and run numbers
            - All time fields (exception time from/to, overtime, bonus, night differential)
            - Job numbers and account information
            - Supervisor authorization details
            
            If you see multiple overtime entries, structure them as separate entries in an 'entries' array.
            Be thorough - extract every piece of overtime information visible on the form.
            
            Return as structured JSON with all fields populated."""
    
    sample_file = genai.upload_file(path=file_path, display_name=os.path.basename(file_path))
    print(f"Uploaded file '{sample_file.display_name}' as: {sample_file.uri}")
    
    # Use more detailed prompt for better extraction
    enhanced_prompt = f"{prompt}\n\nIMPORTANT: This form may contain multiple overtime entries. Extract EVERY single overtime slip you can identify. Look for patterns, repeated sections, or multiple employee entries. If you find multiple overtime slips, structure them in an 'entries' array."
    
    response = gemini_model.generate_content([sample_file, enhanced_prompt])
    print("Gemini extraction response:", response.text)
    return response.text

# Utility to clean and map Gemini output

def clean_and_map_gemini_output(gemini_output, form_type=None):
    cleaned = re.sub(r"^```json|^```|```$", "", gemini_output.strip(), flags=re.MULTILINE).strip()
    import json
    try:
        data = json.loads(cleaned)
        raw_gemini_json = json.dumps(data)  # Store the raw Gemini output
    except Exception as e:
        print("Error parsing cleaned Gemini output:", e)
        return {}, []
    
    # Check if this is a multi-form response (has 'entries' array)
    if isinstance(data, dict) and 'entries' in data and isinstance(data['entries'], list):
        print("Detected multi-form response with entries array")
        forms_data = []
        for i, entry in enumerate(data['entries']):
            print(f"Processing entry {i+1}: {entry}")
            form_data, rows = process_single_form(entry, form_type)
            # Create individual JSON for this form
            individual_json = {
                "form_type": data.get("form_type", "SUPERVISOR'S OVERTIME AUTHORIZATION"),
                "entry": entry
            }
            forms_data.append((form_data, rows, json.dumps(individual_json)))
        return forms_data, raw_gemini_json
    else:
        # Single form response
        print("Processing as single form")
        form_data, rows = process_single_form(data, form_type)
        return [(form_data, rows, raw_gemini_json)], raw_gemini_json

def process_gemini_extraction_hybrid(gemini_output: str, form_type: str = None) -> Tuple[Dict[str, Any], List[Dict[str, Any]], str]:
    """
    Hybrid approach: Store both mapped fields and raw Gemini JSON.
    Can switch between pure extraction and mapped extraction modes.
    """
    cleaned = re.sub(r"^```json|^```|```$", "", gemini_output.strip(), flags=re.MULTILINE).strip()
    
    try:
        raw_data = json.loads(cleaned)
        raw_gemini_json = json.dumps(raw_data, indent=2)
    except Exception as e:
        print("Error parsing cleaned Gemini output:", e)
        return {}, [], ""
    
    if PURE_GEMINI_EXTRACTION:
        # Pure extraction mode - store everything as-is
        return process_pure_extraction(raw_data, form_type, raw_gemini_json)
    else:
        # Mapped extraction mode - use existing logic
        return process_mapped_extraction(raw_data, form_type, raw_gemini_json)

def process_gemini_extraction_dual(gemini_output: str, form_type: str = None) -> Tuple[List[Tuple[Dict[str, Any], List[Dict[str, Any]], str]], str]:
    """
    Process forms in BOTH extraction modes simultaneously.
    Returns a single form with both pure and mapped data stored.
    """
    cleaned = re.sub(r"^```json|^```|```$", "", gemini_output.strip(), flags=re.MULTILINE).strip()
    
    try:
        raw_data = json.loads(cleaned)
        raw_gemini_json = json.dumps(raw_data, indent=2)
    except Exception as e:
        print("Error parsing cleaned Gemini output:", e)
        return [], ""
    
    all_forms = []
    
    # Check if this is a multi-form response
    if isinstance(raw_data, dict) and 'entries' in raw_data and isinstance(raw_data['entries'], list):
        print("Detected multi-form response with entries array")
        
        # Extract employee data from main response for supervisor forms
        employee_data = {}
        if form_type == 'supervisor':
            # Handle both old and new formats
            if 'employee' in raw_data:
                employee_data = raw_data['employee'] or {}
                print(f"Extracted employee data (old format): {employee_data}")
            elif 'employeeDetails' in raw_data:
                employee_data = raw_data['employeeDetails'] or {}
                print(f"Extracted employee data (new format): {employee_data}")
        
        for i, entry in enumerate(raw_data['entries']):
            print(f"Processing entry {i+1}: {entry}")
            
            # Merge employee data with entry data for supervisor forms
            if form_type == 'supervisor' and employee_data:
                merged_entry = {**employee_data, **entry}
                print(f"Merged entry with employee data: {merged_entry}")
            else:
                merged_entry = entry
            
            form_data, rows = process_single_form_combined(merged_entry, form_type, raw_gemini_json)
            individual_json = {
                "form_type": raw_data.get("form_type", "SUPERVISOR'S OVERTIME AUTHORIZATION"),
                "entry": merged_entry
            }
            all_forms.append((form_data, rows, json.dumps(individual_json)))
    else:
        # Single form response
        print("Processing as single form")
        form_data, rows = process_single_form_combined(raw_data, form_type, raw_gemini_json)
        all_forms.append((form_data, rows, raw_gemini_json))
    
    return all_forms, raw_gemini_json

def process_single_form_combined(data: Dict[str, Any], form_type: str, raw_gemini_json: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """
    Process a single form and combine both pure and mapped extraction into one form record.
    """
    # Process pure extraction
    pure_form_data, pure_rows = extract_single_form_pure(data, form_type)
    
    # Process mapped extraction
    mapped_form_data, mapped_rows = process_single_form(data, form_type)
    
    # Combine into a single form record
    combined_form_data = {
        'form_type': form_type,
        'status': 'processed',
        # Store both extraction modes
        'raw_extracted_data_pure': pure_form_data.get('raw_extracted_data', ''),
        'raw_extracted_data_mapped': json.dumps(data),  # Store original raw data for mapped mode
        # Use mapped data as the primary fields (for backward compatibility)
        'pass_number': mapped_form_data.get('pass_number', ''),
        'title': mapped_form_data.get('title', ''),
        'employee_name': mapped_form_data.get('employee_name', ''),
        'rdos': mapped_form_data.get('rdos', ''),
        'actual_ot_date': mapped_form_data.get('actual_ot_date', ''),
        'div': mapped_form_data.get('div', ''),
        'comments': mapped_form_data.get('comments', ''),
        'supervisor_name': mapped_form_data.get('supervisor_name', ''),
        'supervisor_pass_no': mapped_form_data.get('supervisor_pass_no', ''),
        'oto': mapped_form_data.get('oto', ''),
        'oto_amount_saved': mapped_form_data.get('oto_amount_saved', ''),
        'entered_in_uts': mapped_form_data.get('entered_in_uts', ''),
        'regular_assignment': mapped_form_data.get('regular_assignment', ''),
        'report': mapped_form_data.get('report', ''),
        'relief': mapped_form_data.get('relief', ''),
        'todays_date': mapped_form_data.get('todays_date', ''),
        'reg': mapped_form_data.get('reg', ''),
        'superintendent_authorization_signature': mapped_form_data.get('superintendent_authorization_signature', ''),
        'superintendent_authorization_pass': mapped_form_data.get('superintendent_authorization_pass', ''),
        'superintendent_authorization_date': mapped_form_data.get('superintendent_authorization_date', ''),
        'entered_into_uts': mapped_form_data.get('entered_into_uts', ''),
        'overtime_hours': mapped_form_data.get('overtime_hours', ''),
        'report_loc': mapped_form_data.get('report_loc', ''),
        'overtime_location': mapped_form_data.get('overtime_location', ''),
        'report_time': mapped_form_data.get('report_time', ''),
        'relief_time': mapped_form_data.get('relief_time', ''),
        'date_of_overtime': mapped_form_data.get('date_of_overtime', ''),
        'job_number': mapped_form_data.get('job_number', ''),
        'rc_number': mapped_form_data.get('rc_number', ''),
        'acct_number': mapped_form_data.get('acct_number', ''),
        'amount': mapped_form_data.get('amount', ''),
        'reason_rdo': mapped_form_data.get('reason_rdo', False),
        'reason_absentee_coverage': mapped_form_data.get('reason_absentee_coverage', False),
        'reason_no_lunch': mapped_form_data.get('reason_no_lunch', False),
        'reason_early_report': mapped_form_data.get('reason_early_report', False),
        'reason_late_clear': mapped_form_data.get('reason_late_clear', False),
        'reason_save_as_oto': mapped_form_data.get('reason_save_as_oto', False),
        'reason_capital_support_go': mapped_form_data.get('reason_capital_support_go', False),
        'reason_other': mapped_form_data.get('reason_other', False),
        # Store raw JSON for both modes
        'raw_gemini_json': raw_gemini_json,
        'raw_extracted_data': json.dumps(data),  # Store original raw data
        'extraction_mode': 'combined'  # Mark as combined extraction
    }
    
    # Use mapped rows as primary rows
    rows = mapped_rows if mapped_rows else pure_rows
    
    return combined_form_data, rows

def process_pure_extraction(data: Dict[str, Any], form_type: str, raw_gemini_json: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]], str]:
    """
    Pure Gemini extraction - store everything without mapping.
    """
    # Check if this is a multi-form response
    if isinstance(data, dict) and 'entries' in data and isinstance(data['entries'], list):
        print("Detected multi-form response with entries array")
        forms_data = []
        for i, entry in enumerate(data['entries']):
            print(f"Processing entry {i+1}: {entry}")
            form_data, rows = extract_single_form_pure(entry, form_type)
            individual_json = {
                "form_type": data.get("form_type", "SUPERVISOR'S OVERTIME AUTHORIZATION"),
                "entry": entry
            }
            forms_data.append((form_data, rows, json.dumps(individual_json)))
        return forms_data, raw_gemini_json
    else:
        # Single form response
        print("Processing as single form")
        form_data, rows = extract_single_form_pure(data, form_type)
        return [(form_data, rows, raw_gemini_json)], raw_gemini_json

def extract_single_form_pure(data: Dict[str, Any], form_type: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """
    Extract a single form using pure Gemini output without field mapping.
    """
    # Store all top-level fields as-is
    form_data = {
        'form_type': form_type,
        'raw_extracted_data': json.dumps(data),  # Store complete raw data
        'extraction_mode': 'pure'
    }
    
    # Extract any obvious fields that we know about for basic functionality
    # but don't enforce strict mapping
    if 'employee_name' in data:
        form_data['employee_name'] = data['employee_name']
    
    # Flexible pass number extraction
    pass_number = get_flexible_field(data, [
        'pass_number', 'pass', 'PASS', 'pass_no', 'pass no', 'passnumber', 'pass number'
    ])
    if pass_number:
        form_data['pass_number'] = pass_number
    
    if 'title' in data:
        form_data['title'] = data['title']
    if 'comments' in data:
        form_data['comments'] = data['comments']
    
    # Handle rows/entries if they exist
    rows = []
    if 'rows' in data and isinstance(data['rows'], list):
        rows = data['rows']
    elif 'entries' in data and isinstance(data['entries'], list):
        rows = data['entries']
    
    return form_data, rows

def process_mapped_extraction(data: Dict[str, Any], form_type: str, raw_gemini_json: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]], str]:
    """
    Use existing mapped extraction logic but also store raw JSON.
    """
    # Use existing logic but ensure raw JSON is stored
    if isinstance(data, dict) and 'entries' in data and isinstance(data['entries'], list):
        print("Detected multi-form response with entries array")
        forms_data = []
        for i, entry in enumerate(data['entries']):
            print(f"Processing entry {i+1}: {entry}")
            form_data, rows = process_single_form(entry, form_type)
            # Add raw JSON to form data
            form_data['raw_extracted_data'] = json.dumps(entry)
            form_data['extraction_mode'] = 'mapped'
            individual_json = {
                "form_type": data.get("form_type", "SUPERVISOR'S OVERTIME AUTHORIZATION"),
                "entry": entry
            }
            forms_data.append((form_data, rows, json.dumps(individual_json)))
        return forms_data, raw_gemini_json
    else:
        # Single form response
        print("Processing as single form")
        form_data, rows = process_single_form(data, form_type)
        # Add raw JSON to form data
        form_data['raw_extracted_data'] = json.dumps(data)
        form_data['extraction_mode'] = 'mapped'
        return [(form_data, rows, raw_gemini_json)], raw_gemini_json

def process_single_form(data, form_type=None):

    # Normalize keys for mapping
    def normalize_key(k):
        # Handle special cases first
        if k == "RC#":
            return "rc"
        if k == "REPORT LOC.":
            return "report_loc"
        if k == "S/M":
            return "sm"
        
        # General normalization
        normalized = k.lower().replace(" ", "_").replace(".", "").replace("#", "").replace("/", "_").replace("'", "").replace("-", "_")
        return normalized

    # Full slot list for supervisor overtime form and exception claim (hourly) form
    all_fields = [
        # Supervisor/Overtime fields
        "reg_assignment", "reg", "pass_number", "title", "rc_number", "employee_name", "report_loc", "date", "rdos", "date_of_overtime",
        "job_number", "overtime_location", "report_time", "relief_time", "overtime_hours",
        "reason_rdo", "reason_absentee_coverage", "reason_no_lunch", "reason_early_report", "reason_late_clear", "reason_save_as_oto", "reason_capital_support_go", "reason_other",
        "acct_number", "amount",
        "superintendent_authorization_signature", "superintendent_authorization_pass", "superintendent_authorization_date", "entered_into_uts",
        # Exception claim (hourly) fields
        "regular_assignment", "report", "relief", "todays_date", "title", "employee_name", "rdos", "actual_ot_date", "div", "pass_number",
        "exception_code", "line_location", "run_no", "exception_time_from_hh", "exception_time_from_mm", "exception_time_to_hh", "exception_time_to_mm",
        "overtime_hh", "overtime_mm", "ta_job_no", "comments", "oto", "oto_amount_saved_hh", "oto_amount_saved_mm", "entered_in_uts_yes", "entered_in_uts_no", "supervisor_name", "supervisor_pass_no"
    ]
    # Map normalized Gemini keys to our slots
    key_map = {
        # Supervisor/Overtime fields (expanded variants for all possible Gemini outputs)
        "reg": "reg",
        "reg.": "reg",
        "reg_assignment": "regular_assignment",
        "assignment": "pass_number",
        "pass": "pass_number",
        "PASS": "pass_number",  # Direct mapping for original key
        "rc": "rc_number",
        "rc.": "rc_number",
        "rc#": "rc_number",
        "rc_#": "rc_number",
        "rc number": "rc_number",
        "rc ": "rc_number",  # Handle RC with space
        "RC#": "rc_number",  # Direct mapping for original key
        "employee_name": "employee_name",
        "employee name": "employee_name",
        "EMPLOYEE NAME": "employee_name",  # Direct mapping for original key
        "job": "job_number",
        "job.": "job_number",
        "job#": "job_number",
        "job_#": "job_number",
        "job number": "job_number",
        "JOB #": "job_number",  # Direct mapping for original key
        "overtime_location": "overtime_location",
        "overtime location": "overtime_location",
        "OVERTIME LOCATION": "overtime_location",  # Direct mapping for original key
        "report_loc": "report_loc",
        "report_loc.": "report_loc",
        "report_loc..": "report_loc",  # Handle double periods
        "report location": "report_loc",
        "REPORT LOC.": "report_loc",  # Direct mapping for original key
        "report_time": "report_time",
        "report time": "report_time",
        "REPORT TIME": "report_time",  # Direct mapping for original key
        "relief_time": "relief_time",
        "relief time": "relief_time",
        "RELIEF TIME": "relief_time",  # Direct mapping for original key
        "overtime_hours": "overtime_hours",
        "overtime hours": "overtime_hours",
        "OVERTIME HOURS": "overtime_hours",  # Direct mapping for original key
        "date": "todays_date",
        "date_of_overtime": "date_of_overtime",
        "date of overtime": "date_of_overtime",
        "rdo's": "rdos",
        "rdos": "rdos",
        "sm": "rdos",  # Handle S/M field
        "S/M": "rdos",  # Direct mapping for original key
        "reason_for_overtime": "reason_for_overtime",
        "reason for overtime": "reason_for_overtime",
        "REASON FOR OVERTIME": "reason_for_overtime",  # Direct mapping for original key
        "comments": "comments",
        "COMMENTS": "comments",  # Direct mapping for original key
        "acct": "acct_number",
        "acct.": "acct_number",
        "acct#": "acct_number",
        "acct_#": "acct_number",
        "acct number": "acct_number",
        "ACCT #": "acct_number",  # Direct mapping for original key
        "amount": "amount",
        "supervisors_signature": "superintendent_authorization_signature",
        "supervisor's_signature": "superintendent_authorization_signature",
        "superintendent_authorization_signature": "superintendent_authorization_signature",
        "superintendent's_authorization_-_signature": "superintendent_authorization_signature",
        "superintendent_authorization_pass": "superintendent_authorization_pass",
        "superintendent's_authorization_-_pass": "superintendent_authorization_pass",
        "superintendent_authorization_date": "superintendent_authorization_date",
        "superintendent's_authorization_-_date": "superintendent_authorization_date",
        "entered_into_uts": "entered_into_uts",
        # Additional mappings for specific Gemini output formats
        "reg_assignment": "regular_assignment",
        "report_loc": "report_loc",
        "report loc": "report_loc",
        "report location": "report_loc",
        "overtime_location": "overtime_location",
        "overtime location": "overtime_location",
        "report_time": "report_time",
        "report time": "report_time",
        "relief_time": "relief_time",
        "relief time": "relief_time",
        "overtime_hours": "overtime_hours",
        "overtime hours": "overtime_hours",
        "rdo's": "rdos",
        "rdos": "rdos",
        "reason_for_overtime": "reason_for_overtime",
        "reason for overtime": "reason_for_overtime",
        "acct": "acct_number",
        "acct#": "acct_number",
        "acct number": "acct_number",
        "superintendent's_authorization": "superintendent_authorization",
        # Exception claim (hourly) fields
        "regular_assignment": "regular_assignment",
        "report": "report",
        "relief": "relief",
        "todays_date": "todays_date",
        "today's_date": "todays_date",
        "pass_number": "pass_number",
        "title": "title",
        "TITLE": "title",  # Direct mapping for original key
        "employee_name": "employee_name",
        "rdos": "rdos",
        "actual_ot_date": "actual_ot_date",
        "div": "div",
        "code": "exception_code",
        "line_location": "line_location",
        "run_no": "run_no",
        "exception_time_from_hh": "exception_time_from_hh",
        "exception_time_from_mm": "exception_time_from_mm",
        "exception_time_to_hh": "exception_time_to_hh",
        "exception_time_to_mm": "exception_time_to_mm",
        "overtime_hh": "overtime_hh",
        "overtime_mm": "overtime_mm",
        "ta_job_no": "ta_job_no",
        "comments": "comments",
        "oto": "oto",
        "oto_amount_saved_hh": "oto_amount_saved_hh",
        "oto_amount_saved_mm": "oto_amount_saved_mm",
        "entered_in_uts_yes": "entered_in_uts_yes",
        "entered_in_uts_no": "entered_in_uts_no",
        "supv_name": "supervisor_name",
        "supervisor_name": "supervisor_name",
        "pass_no": "supervisor_pass_no",
        "supervisor_pass_no": "supervisor_pass_no",
        "job_": "job_number",
        "rc_": "rc_number",
        "acct_": "acct_number",
        "report_time": "report_time",
        "relief_time": "relief_time",
        "overtime_hours": "overtime_hours",
        "superintendent_s_authorization___pass": "superintendent_authorization_pass",
        "superintendent_s_authorization___date": "superintendent_authorization_date",
        "superintendent_s_authorization___signature": "superintendent_authorization_signature",
        "entered_into_uts": "entered_into_uts"
    }
    # Robust time/location field variants
    key_map.update({
        "report_time": "report_time",
        "report time": "report_time",
        "report-time": "report_time",
        "report.time": "report_time",
        "reporttime": "report_time",
        "relief_time": "relief_time",
        "relief time": "relief_time",
        "relief-time": "relief_time",
        "relief.time": "relief_time",
        "relieftime": "relief_time",
        "report_loc": "report_loc",
        "report loc": "report_loc",
        "report-loc": "report_loc",
        "report.loc": "report_loc",
        "reportloc": "report_loc",
        "overtime_location": "overtime_location",
        "overtime location": "overtime_location",
        "overtime-location": "overtime_location",
        "overtime.location": "overtime_location",
        "overtimelocation": "overtime_location",
    })
    # Checkbox mapping
    checkbox_map = {
        "rdo": "reason_rdo",
        "absentee_coverage": "reason_absentee_coverage",
        "no_lunch": "reason_no_lunch",
        "early_report": "reason_early_report",
        "late_clear": "reason_late_clear",
        "save_as_oto": "reason_save_as_oto",
        "capital_support_go": "reason_capital_support_go",
        "other": "reason_other"
    }
    form_data = {field: False if field.startswith('reason_') else '' for field in all_fields}
    row = {}
    # Flatten nested dicts if needed
    def flatten(d, parent_key=""):
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}_{k}" if parent_key else k
            if isinstance(v, dict):
                # Special handling for supervisor nested fields
                if parent_key.lower().startswith("superintendent's_authorization") or k.lower().startswith("superintendent's_authorization"):
                    # Map nested keys directly to their backend fields
                    for subk, subv in v.items():
                        subkey_norm = normalize_key(f"superintendent's_authorization_-_{subk}")
                        items.append((subkey_norm, subv))
                elif k.lower() == "superintendent's_authorization":
                    # Handle the nested superintendent authorization object
                    for subk, subv in v.items():
                        subkey_norm = normalize_key(f"superintendent's_authorization_-_{subk}")
                        items.append((subkey_norm, subv))
                else:
                    items.extend(flatten(v, new_key).items())
            else:
                items.append((new_key, v))
        return dict(items)
    flat_data = flatten(data)
    print("FLATTENED GEMINI DATA:", flat_data)
    # Map fields
    for k, v in flat_data.items():
        norm_k = normalize_key(k)
        print(f"Mapping: '{k}' -> '{norm_k}' = '{v}'")
        
        # Handle nested employee data first
        if k == "employee_pass_number" or k == "employee_pass":
            form_data["pass_number"] = v
        elif k == "employee_name" or k == "employee_name":
            form_data["employee_name"] = v
        elif k == "employee_title" or k == "employee_title":
            form_data["title"] = v
        elif k == "employee_rdoe" or k == "employee_rdo":
            form_data["rdos"] = v
        elif k == "employee_rc_number" or k == "employee_rc":
            form_data["rc_number"] = v
        
        # Handle direct field mappings for merged data
        elif k == "name":
            form_data["employee_name"] = v
        elif k == "rduos" or k == "rdo":
            form_data["rdos"] = v
        
        # Handle new camelCase field names
        elif k == "employeeName":
            form_data["employee_name"] = v
        elif k == "passNumber":
            form_data["pass_number"] = v
        elif k == "rDOs":
            form_data["rdos"] = v
        elif k == "rcNumber":
            form_data["rc_number"] = v
        elif k == "jobNumber":
            form_data["job_number"] = v
        elif k == "overtimeLocation":
            form_data["overtime_location"] = v
        elif k == "reportTime":
            form_data["report_time"] = v
        elif k == "reliefTime":
            form_data["relief_time"] = v
        elif k == "overtimeHours":
            form_data["overtime_hours"] = v
        elif k == "accountNumber":
            form_data["acct_number"] = v
        elif k == "reportLocation":
            form_data["report_loc"] = v
        
        # Handle various date field formats
        elif k in ["DATE", "date", "s_m", "W/T", "w_t"]:
            # Map various date fields to date_of_overtime
            if v and str(v).strip() and str(v).strip() != "None":
                form_data["date_of_overtime"] = str(v).strip()
                print(f"Mapped date field '{k}' -> date_of_overtime: {v}")
        
        # Handle assignment/job number variations
        elif k in ["ASSIGNMENT", "assignment"]:
            if v and str(v).strip() and str(v).strip() != "None":
                form_data["pass_number"] = str(v).strip()
                print(f"Mapped assignment field '{k}' -> pass_number: {v}")
        elif k in ["RBG", "rbg"]:
            if v and str(v).strip() and str(v).strip() != "None":
                form_data["job_number"] = str(v).strip()
                print(f"Mapped RBG field '{k}' -> job_number: {v}")
        
        # Handle reason_for_overtime as string or list
        elif norm_k == "reason_for_overtime":
            if isinstance(v, str):
                v = [v]
            for reason in v:
                reason_norm = normalize_key(reason)
                if reason_norm in checkbox_map:
                    form_data[checkbox_map[reason_norm]] = True
        # Handle reason field mapping for supervisor forms
        elif norm_k == "reason":
            # Map reason to appropriate checkbox
            reason_lower = str(v).lower()
            if "rdo" in reason_lower:
                form_data["reason_rdo"] = True
            elif "absentee" in reason_lower or "coverage" in reason_lower:
                form_data["reason_absentee_coverage"] = True
            elif "lunch" in reason_lower:
                form_data["reason_no_lunch"] = True
            elif "early" in reason_lower and "report" in reason_lower:
                form_data["reason_early_report"] = True
            elif "late" in reason_lower and "clear" in reason_lower:
                form_data["reason_late_clear"] = True
            elif "oto" in reason_lower:
                form_data["reason_save_as_oto"] = True
            elif "capital" in reason_lower or "support" in reason_lower:
                form_data["reason_capital_support_go"] = True
            else:
                form_data["reason_other"] = True
        
        # Handle new reasonForOvertime object format
        elif k == "reasonForOvertime" and isinstance(v, dict):
            # Map boolean values from reasonForOvertime object
            if v.get('rdo'):
                form_data["reason_rdo"] = True
            if v.get('absenteeCoverage'):
                form_data["reason_absentee_coverage"] = True
            if v.get('noLunch'):
                form_data["reason_no_lunch"] = True
            if v.get('earlyReport'):
                form_data["reason_early_report"] = True
            if v.get('lateClear'):
                form_data["reason_late_clear"] = True
            if v.get('saveAsOto'):
                form_data["reason_save_as_oto"] = True
            if v.get('capitalSupportGo'):
                form_data["reason_capital_support_go"] = True
            if v.get('other'):
                form_data["reason_other"] = True
        # Custom logic for ambiguous keys
        elif norm_k == "pass":
            if "superintendent_authorization" in k.lower():
                form_data["superintendent_authorization_pass"] = v
            else:
                form_data["pass_number"] = v
        elif norm_k == "date":
            if "superintendent_authorization" in k.lower():
                form_data["superintendent_authorization_date"] = v
            elif "date_of_overtime" in k.lower():
                form_data["date_of_overtime"] = v
            elif form_type == 'supervisor':
                form_data["todays_date"] = v
            # else: do not map 'date' for hourly forms
        elif norm_k == "job":
            form_data["job_number"] = v
        elif norm_k == "rc":
            form_data["rc_number"] = v
        elif norm_k == "rc_number":
            form_data["rc_number"] = v
        elif k in ["RC#", "RC#", "rc#"]:
            form_data["rc_number"] = v
        elif norm_k == "report_loc":
            form_data["report_loc"] = v
        elif norm_k == "reg_assignment":
            form_data["regular_assignment"] = v
        elif norm_k == "overtime_location":
            form_data["overtime_location"] = v
        elif norm_k == "report_time":
            form_data["report_time"] = v
        elif norm_k == "relief_time":
            form_data["relief_time"] = v
        elif norm_k == "overtime_hours":
            form_data["overtime_hours"] = v
        elif norm_k == "rdos":
            form_data["rdos"] = v
        elif norm_k == "sm":
            form_data["rdos"] = v
        elif norm_k == "acct":
            form_data["acct_number"] = v
        elif norm_k == "account_number":
            form_data["acct_number"] = v
        elif norm_k == "report_location":
            form_data["report_loc"] = v
        elif norm_k == "superintendents_authorization":
            # Handle combined superintendent authorization field (e.g., "713026 07/06/25")
            if isinstance(v, str) and ' ' in v:
                parts = v.split(' ', 1)  # Split on first space
                if len(parts) == 2:
                    form_data["superintendent_authorization_pass"] = parts[0]
                    form_data["superintendent_authorization_date"] = parts[1]
                else:
                    form_data["superintendent_authorization_pass"] = v
            else:
                form_data["superintendent_authorization_pass"] = v
        elif norm_k in key_map:
            form_data[key_map[norm_k]] = v
        elif norm_k in checkbox_map:
            form_data[checkbox_map[norm_k]] = bool(v)
        elif norm_k.startswith("reason_for_overtime_"):
            # e.g. reason_for_overtime_rdo: True
            reason = norm_k.replace("reason_for_overtime_", "")
            if reason in checkbox_map:
                form_data[checkbox_map[reason]] = bool(v)
        else:
            print(f"Unmapped Gemini key: {k} -> {v}")
    # Clean the form data to handle None values and lists
    for key, value in form_data.items():
        if value is None:
            form_data[key] = ''
        elif isinstance(value, list):
            form_data[key] = ', '.join(str(item) for item in value if item is not None)
        elif isinstance(value, bool):
            form_data[key] = 1 if value else 0
        elif key.startswith('reason_') and isinstance(value, str):
            # Convert string boolean values to integers for reason fields
            form_data[key] = 1 if value.lower() in ['true', '1', 'yes', 'on'] else 0
        else:
            form_data[key] = str(value) if value is not None else ''
    print("FINAL FORM DATA (mapped):", form_data)
    # Log dashboard-relevant fields
    dashboard_fields = ['overtime_hours', 'job_number', 'title', 'report_loc', 'overtime_location']
    for field in dashboard_fields:
        print(f"DASHBOARD FIELD {field}: {form_data.get(field)}")

    # --- PATCH: Ensure dashboard-relevant mapped fields are always set using flexible lookup from original Gemini data ---
    dashboard_fields = {
        'overtime_hours': [
            'overtime_hours', 'overtime', 'hours', 'ot_hours', 'ot', 'total_overtime', 'overtime total', 'ot total', 'ot time', 'total ot', 'total_ot', 'ot time (hh:mm)', 'overtime (hh:mm)'
        ],
        'job_number': [
            'job_number', 'job no', 'job_no', 'job', 'job#', 'job number', 'ta_job_no', 'ta job no', 'jobnum', 'jobnumber', 'job id', 'jobid'
        ],
        'title': [
            'title', 'position', 'job_title', 'role', 'job title', 'employee_title', 'employee title'
        ],
        'report_loc': [
            'report_loc', 'report location', 'location', 'reportloc', 'report', 'loc', 'report station', 'station', 'reporting location'
        ],
        'overtime_location': [
            'overtime_location', 'overtime location', 'overtimelocation', 'ot_location', 'otloc', 'location', 'ot location', 'ot station', 'overtime station'
        ]
    }
    for field, variants in dashboard_fields.items():
        if not form_data.get(field):
            value = get_flexible_field(data, variants)
            if value:
                form_data[field] = value
    # Log the final values for dashboard fields
    print("DASHBOARD FIELD VALUES:")
    for field in dashboard_fields:
        print(f"  {field}: {form_data.get(field)}")
    
    # Debug: Show the final form_data after all processing
    print("DEBUG: Final form_data after dashboard field lookup:")
    for key, value in form_data.items():
        if value and key in ['pass_number', 'employee_name', 'title', 'job_number', 'rc_number', 'overtime_hours', 'report_loc', 'overtime_location']:
            print(f"  {key}: {value}")

    # For exception claim (hourly) forms, build a row if relevant fields are present
    row_fields = [
        "exception_code", "line_location", "run_no",
        "exception_time_from_hh", "exception_time_from_mm",
        "exception_time_to_hh", "exception_time_to_mm",
        "overtime_hh", "overtime_mm",
        "bonus_hh", "bonus_mm",
        "nite_diff_hh", "nite_diff_mm",
        "ta_job_no"
    ]
    row = {field: form_data.get(field, '') for field in row_fields}
    # Add a row if ANY of the key fields are present (not just code/location/run_no)
    if any(str(v).strip() for v in row.values()):
        rows = [row]
    else:
        rows = []
    # Debug: Show what we're returning from process_single_form
    print("DEBUG: Returning from process_single_form:")
    for key, value in form_data.items():
        if value and key in ['pass_number', 'employee_name', 'title', 'job_number', 'rc_number', 'overtime_hours', 'report_loc', 'overtime_location']:
            print(f"  {key}: {value}")
    
    return form_data, rows

def is_duplicate_form(form_data, form_type):
    """
    Check if a form is a duplicate based on key fields.
    Returns True if duplicate found, False otherwise.
    """
    try:
        import sqlite3
        
        # For Pure Extraction mode, be very lenient - only check for exact duplicates
        if PURE_GEMINI_EXTRACTION:
            print("Pure Extraction mode: Skipping strict duplicate detection")
            return False
        
        # Key fields for duplicate detection
        pass_number = form_data.get('pass_number', '')
        employee_name = form_data.get('employee_name', '')
        overtime_hours = form_data.get('overtime_hours', '')
        date_of_overtime = form_data.get('date_of_overtime', '')
        job_number = form_data.get('job_number', '')
        
        # If we don't have the key fields, we can't check for duplicates
        if not pass_number or not overtime_hours or not date_of_overtime:
            print(f"Cannot check for duplicates - missing key fields: pass_number={pass_number}, overtime_hours={overtime_hours}, date_of_overtime={date_of_overtime}")
            return False
        
        with sqlite3.connect('forms.db', timeout=10) as conn:
            c = conn.cursor()
            
            # Check for existing forms with same key fields
            query = '''
                SELECT id, employee_name, overtime_hours, date_of_overtime, job_number 
                FROM exception_forms 
                WHERE form_type = ? 
                AND pass_number = ? 
                AND overtime_hours = ? 
                AND date_of_overtime = ?
            '''
            
            c.execute(query, (form_type, pass_number, overtime_hours, date_of_overtime))
            existing_forms = c.fetchall()
            
            if existing_forms:
                print(f"Duplicate detected: {len(existing_forms)} existing form(s) with same pass_number={pass_number}, overtime_hours={overtime_hours}, date_of_overtime={date_of_overtime}")
                for existing in existing_forms:
                    print(f"  Existing form ID {existing[0]}: {existing[1]} - {existing[2]} hours on {existing[3]} (Job: {existing[4]})")
                return True
            
            # Additional check: if job_number is also the same, it's definitely a duplicate
            if job_number:
                query_with_job = '''
                    SELECT id, employee_name, overtime_hours, date_of_overtime, job_number 
                    FROM exception_forms 
                    WHERE form_type = ? 
                    AND pass_number = ? 
                    AND overtime_hours = ? 
                    AND date_of_overtime = ? 
                    AND job_number = ?
                '''
                
                c.execute(query_with_job, (form_type, pass_number, overtime_hours, date_of_overtime, job_number))
                existing_with_job = c.fetchall()
                
                if existing_with_job:
                    print(f"Duplicate detected (with job number): {len(existing_with_job)} existing form(s) with same pass_number={pass_number}, overtime_hours={overtime_hours}, date_of_overtime={date_of_overtime}, job_number={job_number}")
                    return True
            
            print(f"No duplicates found for: {employee_name} - {overtime_hours} hours on {date_of_overtime}")
            return False
            
    except Exception as e:
        print(f"Error checking for duplicates: {e}")
        # If we can't check for duplicates, allow the form to be processed
        return False

def is_blank_or_crossed_out(image_path):
    # Enhanced blank/crossed-out detection: be less aggressive to capture more forms
    try:
        from PIL import Image
        img = Image.open(image_path).convert('L')
        width, height = img.size
        
        # For Pure Extraction mode, be extremely lenient - process almost everything
        if PURE_GEMINI_EXTRACTION:
            # Only skip if completely blank
            nonwhite = sum(1 for p in img.getdata() if p < 240)
            if nonwhite < 100:  # Very low threshold for pure extraction
                print(f"Pure Extraction: Segment appears completely blank (only {nonwhite} non-white pixels)")
                return True
            print(f"Pure Extraction: Processing segment with {nonwhite} non-white pixels")
            return False
        
        # Regular mode: Enhanced blank/crossed-out detection
        nonwhite = sum(1 for p in img.getdata() if p < 240)
        
        # More lenient threshold - only skip if almost completely blank
        # This allows partial forms and form fragments to be processed
        if nonwhite < 500:  # Reduced from 1000 to 500
            print(f"Segment appears blank (only {nonwhite} non-white pixels)")
            return True
            
        # Additional check: if segment is very small, don't skip it
        if width < 200 or height < 200:
            print(f"Segment is small ({width}x{height}) but may contain valuable data")
            return False
            
        print(f"Segment has {nonwhite} non-white pixels - processing")
        return False
        
    except Exception as e:
        print(f"Error in blank/crossed-out detection: {e}")
        # If we can't determine, process the segment anyway
        return False

# Flexible field lookup for raw JSON

def get_flexible_field(data, possible_keys):
    """
    Search for a value in a dict by a list of possible field names (case-insensitive, with normalization).
    Handles flat dicts only. For nested dicts, extend as needed.
    """
    def normalize(s):
        return s.lower().replace(' ', '').replace('_', '').replace('-', '')
    norm_data = {normalize(k): v for k, v in data.items()}
    for key in possible_keys:
        norm_key = normalize(key)
        if norm_key in norm_data:
            return norm_data[norm_key]
    return None

# --- PATCH: Set file_name using flexible lookup for both mapped and pure extraction modes ---
def get_flexible_file_name(form_data, raw_json=None, fallback=None):
    file_name = form_data.get('pass_number')
    if not file_name and raw_json:
        file_name = (
            raw_json.get('pass_number') or
            raw_json.get('file_name') or
            raw_json.get('filename') or
            raw_json.get('name')
        )
    if not file_name and fallback:
        file_name = fallback
    return str(file_name) if file_name else 'N/A'

@app.route('/upload/hourly', methods=['POST'])
def upload_hourly_file():
    return handle_upload(form_type='hourly')

@app.route('/upload/supervisor', methods=['POST'])
def upload_supervisor_file():
    return handle_upload(form_type='supervisor')

@app.route('/cleanup-duplicates', methods=['POST'])
def cleanup_duplicates():
    """
    Clean up duplicate forms in the database.
    Keeps the first occurrence and removes duplicates.
    """
    try:
        import sqlite3
        
        with sqlite3.connect('forms.db', timeout=10) as conn:
            c = conn.cursor()
            
            # Find and remove duplicates based on key fields
            cleanup_query = '''
                DELETE FROM exception_forms 
                WHERE id NOT IN (
                    SELECT MIN(id) 
                    FROM exception_forms 
                    WHERE form_type = 'supervisor'
                    GROUP BY pass_number, overtime_hours, date_of_overtime, job_number
                    HAVING pass_number IS NOT NULL 
                    AND pass_number != '' 
                    AND overtime_hours IS NOT NULL 
                    AND overtime_hours != ''
                    AND date_of_overtime IS NOT NULL 
                    AND date_of_overtime != ''
                )
                AND form_type = 'supervisor'
            '''
            
            # Get count before cleanup
            c.execute("SELECT COUNT(*) FROM exception_forms WHERE form_type='supervisor'")
            count_before = c.fetchone()[0]
            
            # Execute cleanup
            c.execute(cleanup_query)
            deleted_count = c.rowcount
            
            # Get count after cleanup
            c.execute("SELECT COUNT(*) FROM exception_forms WHERE form_type='supervisor'")
            count_after = c.fetchone()[0]
            
            conn.commit()
            
            return jsonify({
                'message': 'Duplicate cleanup completed',
                'deleted_count': deleted_count,
                'count_before': count_before,
                'count_after': count_after
            }), 200
            
    except Exception as e:
        return jsonify({'error': f'Cleanup failed: {str(e)}'}), 500

@app.route('/enable-pure-extraction', methods=['POST'])
def enable_pure_extraction():
    """
    Enable Pure Extraction mode to capture everything from PDFs.
    """
    try:
        global PURE_GEMINI_EXTRACTION
        PURE_GEMINI_EXTRACTION = True
        
        return jsonify({
            'message': 'Pure Extraction mode enabled',
            'mode': 'pure',
            'description': 'System will now capture everything from PDF segments without duplicate detection or strict field validation'
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to enable Pure Extraction: {str(e)}'}), 500

@app.route('/disable-pure-extraction', methods=['POST'])
def disable_pure_extraction():
    """
    Disable Pure Extraction mode and return to normal mapped extraction.
    """
    try:
        global PURE_GEMINI_EXTRACTION
        PURE_GEMINI_EXTRACTION = False
        
        return jsonify({
            'message': 'Pure Extraction mode disabled',
            'mode': 'mapped',
            'description': 'System will now use normal mapped extraction with duplicate detection and field validation'
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to disable Pure Extraction: {str(e)}'}), 500

# Refactor the upload logic into a helper

def handle_upload(form_type):
    try:
        files = request.files.getlist('files')
        if not files:
            # fallback for single file upload (old clients)
            if 'file' in request.files:
                files = [request.files['file']]
            else:
                return jsonify({'error': 'No file(s) part in the request.'}), 400

        from db import store_exception_form, log_audit
        username = request.form.get('username') or request.args.get('username') or (request.json.get('username') if request.is_json else None)
        if not username:
            username = 'unknown'
        success = 0
        failed = 0
        for file in files:
            if file.filename == '':
                failed += 1
                continue
            try:
                # Save files in subfolders by form_type
                subfolder = form_type if form_type in ['hourly', 'supervisor'] else ''
                target_folder = os.path.join(UPLOAD_FOLDER, subfolder) if subfolder else UPLOAD_FOLDER
                os.makedirs(target_folder, exist_ok=True)
                filepath = os.path.join(target_folder, file.filename)
                file.save(filepath)

                # If supervisor and PDF, use MAXIMUM segmentation to extract ALL possible forms
                if form_type == 'supervisor' and file.filename.lower().endswith('.pdf'):
                    print(f"Processing PDF with MAXIMUM extraction: {file.filename}")
                    with pdfplumber.open(filepath) as pdf:
                        total_pages = len(pdf.pages)
                        print(f"PDF has {total_pages} pages")
                        
                        for i, page in enumerate(pdf.pages):
                            print(f"Processing page {i+1}/{total_pages}")
                            img = page.to_image(resolution=300).original
                            width, height = img.size
                            print(f"Page {i+1} dimensions: {width}x{height}")
                            
                            # MAXIMUM segmentation: Extract every possible form region
                            segments = []
                            
                            # Method 1: Standard halves (top/bottom)
                            top_half = img.crop((0, 0, width, height // 2))
                            bottom_half = img.crop((0, height // 2, width, height))
                            segments.extend([top_half, bottom_half])
                            
                            # Method 2: Quarters for dense pages
                            if height > 1000:
                                quarter_height = height // 4
                                segments.extend([
                                    img.crop((0, 0, width, quarter_height)),
                                    img.crop((0, quarter_height, width, quarter_height * 2)),
                                    img.crop((0, quarter_height * 2, width, quarter_height * 3)),
                                    img.crop((0, quarter_height * 3, width, height))
                                ])
                            
                            # Method 3: Eighths for very dense pages
                            if height > 1500:
                                eighth_height = height // 8
                                for e in range(8):
                                    y_start = e * eighth_height
                                    y_end = (e + 1) * eighth_height if e < 7 else height
                                    segments.append(img.crop((0, y_start, width, y_end)))
                            
                            # Method 4: Tenths for maximum coverage
                            if height > 2000:
                                tenth_height = height // 10
                                for t in range(10):
                                    y_start = t * tenth_height
                                    y_end = (t + 1) * tenth_height if t < 9 else height
                                    segments.append(img.crop((0, y_start, width, y_end)))
                            
                            # Method 5: Dynamic segmentation based on content density
                            try:
                                gray_img = img.convert('L')
                                # More aggressive segmentation - check every 1/8 of the page
                                for y in range(0, height, height // 8):
                                    if y + height // 8 < height:
                                        segment = img.crop((0, y, width, y + height // 8))
                                        segments.append(segment)
                                
                                # Additional segments for overlapping coverage
                                for y in range(height // 16, height, height // 8):
                                    if y + height // 8 < height:
                                        segment = img.crop((0, y, width, y + height // 8))
                                        segments.append(segment)
                            except Exception as e:
                                print(f"Dynamic segmentation failed for page {i+1}: {e}")
                                pass
                            
                            # Method 6: Full page as single segment (for forms that span entire page)
                            segments.append(img)
                            
                            print(f"Page {i+1} generated {len(segments)} potential segments")
                            
                            # Enhanced duplicate detection and processing
                            unique_segments = []
                            processed_segments = 0
                            
                            for segment in segments:
                                # Check if segment is too small or empty
                                if segment.size[0] > 100 and segment.size[1] > 100:
                                    # Convert to bytes for comparison
                                    segment_bytes = io.BytesIO()
                                    segment.save(segment_bytes, format='PNG')
                                    segment_data = segment_bytes.getvalue()
                                    
                                    # Enhanced duplicate detection with similarity threshold
                                    is_duplicate = False
                                    for existing in unique_segments:
                                        if _segment_similarity(segment_data, existing):
                                            is_duplicate = True
                                            break
                                    
                                    if not is_duplicate:
                                        unique_segments.append(segment_data)
                                        processed_segments += 1
                            
                            print(f"Page {i+1} has {processed_segments} unique segments to process")
                            
                            # Process each unique segment with enhanced error handling
                            for j, segment_data in enumerate(unique_segments):
                                segment_img = Image.open(io.BytesIO(segment_data))
                                segment_path = os.path.join(target_folder, f"{os.path.splitext(file.filename)[0]}_page{i+1}_segment{j+1}.png")
                                segment_img.save(segment_path)
                                
                                if is_blank_or_crossed_out(segment_path):
                                    print(f"Skipped blank/crossed-out segment: {segment_path}")
                                    continue
                                
                                # Use Gemini to extract details with dual approach (both pure and mapped)
                                print(f"Processing segment {j+1}/{len(unique_segments)} from page {i+1}")
                                try:
                                    gemini_output = gemini_extract_file_details(segment_path, form_type=form_type)
                                    if gemini_output:
                                        print(f"--- Gemini Output for segment {j+1} ---")
                                        print(gemini_output[:500] + "..." if len(gemini_output) > 500 else gemini_output)
                                        print("--- END Gemini Output ---")
                                        
                                        forms_data, raw_gemini_json = process_gemini_extraction_dual(gemini_output, form_type=form_type)
                                        print(f"Extracted {len(forms_data)} forms from segment {j+1}")
                                    else:
                                        print(f"No Gemini output for segment {j+1} - skipping")
                                        forms_data, raw_gemini_json = [], ''
                                        continue
                                except Exception as e:
                                    print(f"Error processing segment {j+1}: {e}")
                                    forms_data, raw_gemini_json = [], ''
                                    continue
                                
                                # Process each form from the response with deduplication
                                for form_data, rows, individual_json in forms_data:
                                    # Use the individual_json for the raw_gemini_json
                                    form_data['raw_gemini_json'] = individual_json
                                    # --- PATCH: Set file_name using flexible lookup for both mapped and pure extraction modes ---
                                    form_data['file_name'] = get_flexible_file_name(form_data, raw_gemini_json, os.path.basename(segment_path))
                                    if form_data:
                                        # For Pure Extraction mode, be more lenient with required fields
                                        if PURE_GEMINI_EXTRACTION:
                                            required_form_fields = [
                                                'status', 'file_name'  # Only absolutely essential fields
                                            ]
                                        else:
                                            required_form_fields = [
                                                'pass_number', 'title', 'employee_name', 'rdos', 'actual_ot_date', 'div',
                                                'comments', 'supervisor_name', 'supervisor_pass_no', 'oto', 'oto_amount_saved',
                                                'entered_in_uts', 'regular_assignment', 'report', 'relief', 'todays_date', 'status', 'file_name'
                                            ]
                                        for key in required_form_fields:
                                            if key not in form_data:
                                                form_data[key] = 'N/A'
                                        form_data['status'] = 'processed'
                                        
                                        # Check for duplicates based on overtime hours and date
                                        # SKIP duplicate detection for Pure Extraction mode - we want everything!
                                        if not PURE_GEMINI_EXTRACTION and is_duplicate_form(form_data, form_type):
                                            print(f"Skipping duplicate form: {form_data.get('employee_name', 'Unknown')} - {form_data.get('overtime_hours', 'Unknown hours')} on {form_data.get('date_of_overtime', 'Unknown date')}")
                                            continue
                                        
                                    else:
                                        form_data = {key: '' for key in required_form_fields}
                                        form_data['file_name'] = os.path.basename(img_path)
                                        form_data['comments'] = f"Gemini extraction failed for {img_path}. Output: {gemini_output if gemini_output else 'None'}"
                                        form_data['status'] = 'error'
                                        if not rows:
                                            rows = []
                                    upload_date = datetime.datetime.now().isoformat()
                                    form_id = store_exception_form(form_data, rows, username, form_type=form_type, upload_date=upload_date)
                                    if not form_id:
                                        failed += 1
                                        continue
                                    # Safely get pass_number for audit log
                                    pass_number = 'N/A'
                                    if isinstance(form_data, dict):
                                        pass_number = form_data.get('pass_number', 'N/A')
                                    log_audit(username, 'upload', 'form', form_id, f"Form uploaded: {pass_number}")
                                    success += 1
                    continue  # Skip the rest of the loop for supervisor PDFs

                # Enhanced processing for hourly forms to extract maximum overtime slips
                if form_type == 'hourly':
                    # Try to detect multiple forms in the document first
                    multiple_forms = detect_multiple_forms_in_document(filepath, form_type)
                    
                    if multiple_forms:
                        print(f"Detected {len(multiple_forms)} potential form regions in hourly document")
                        # Process each detected region
                        all_forms_data = []
                        for i, segment in enumerate(multiple_forms):
                            segment_path = os.path.join(target_folder, f"{os.path.splitext(file.filename)[0]}_segment{i+1}.png")
                            segment.save(segment_path)
                            
                            if is_blank_or_crossed_out(segment_path):
                                print(f"Skipped blank segment {i+1}")
                                continue
                            
                            # Extract from this segment
                            segment_output = gemini_extract_file_details(segment_path, form_type=form_type)
                            if segment_output:
                                segment_forms, _ = process_gemini_extraction_dual(segment_output, form_type=form_type)
                                all_forms_data.extend(segment_forms)
                        
                        # Combine all extracted forms
                        if all_forms_data:
                            gemini_output = json.dumps({"entries": [form[0] for form in all_forms_data]})
                            print(f"Successfully extracted {len(all_forms_data)} forms from multiple regions")
                        else:
                            gemini_output = None
                    else:
                        # Standard extraction with enhanced prompts
                        gemini_output = gemini_extract_file_details(filepath, form_type=form_type)
                        
                        # If no multiple entries found, try with enhanced prompt
                        if gemini_output and 'entries' not in gemini_output:
                            enhanced_prompt = """This is an hourly employee overtime form. Look VERY carefully for multiple overtime entries.
                            
                            CRITICAL: These forms often contain multiple overtime slips for the same employee on different dates or times.
                            Look for:
                            - Multiple date entries
                            - Multiple time ranges
                            - Multiple exception codes
                            - Multiple line locations or run numbers
                            - Any repeated patterns that suggest multiple overtime entries
                            
                            If you find ANY indication of multiple overtime entries, structure them in an 'entries' array.
                            Be extremely thorough - these forms are designed to capture multiple overtime instances."""
                            
                            enhanced_output = gemini_extract_file_details(filepath, enhanced_prompt, form_type=form_type)
                            if enhanced_output and 'entries' in enhanced_output:
                                gemini_output = enhanced_output
                                print("Enhanced extraction found multiple overtime entries!")
                else:
                    # Default: process as a single file (for non-PDF supervisor uploads)
                    gemini_output = gemini_extract_file_details(filepath, form_type=form_type)
                print("--- Gemini Output ---")
                print(gemini_output)
                print("--- END Gemini Output ---")
                forms_data, raw_gemini_json = process_gemini_extraction_dual(gemini_output, form_type=form_type) if gemini_output else ([], '')
                
                # Process each form from the response
                for form_data, rows, individual_json in forms_data:
                    # Use the individual_json for the raw_gemini_json
                    form_data['raw_gemini_json'] = individual_json
                    # --- PATCH: Set file_name using flexible lookup for both mapped and pure extraction modes ---
                    form_data['file_name'] = get_flexible_file_name(form_data, raw_gemini_json, file.filename)
                    if form_data:
                        required_form_fields = [
                            'pass_number', 'title', 'employee_name', 'rdos', 'actual_ot_date', 'div',
                            'comments', 'supervisor_name', 'supervisor_pass_no', 'oto', 'oto_amount_saved',
                            'entered_in_uts', 'regular_assignment', 'report', 'relief', 'todays_date', 'status', 'file_name'
                        ]
                        for key in required_form_fields:
                            if key not in form_data:
                                form_data[key] = 'N/A'
                        form_data['status'] = 'processed'
                    else:
                        form_data = {key: '' for key in required_form_fields}
                        form_data['file_name'] = file.filename
                        form_data['comments'] = f"Gemini extraction failed for {file.filename}. Output: {gemini_output if gemini_output else 'None'}"
                        form_data['status'] = 'error'
                        if not rows:
                            rows = []
                    upload_date = datetime.datetime.now().isoformat()
                    print(f"DEBUG: About to store form. form_data type: {type(form_data)}, rows type: {type(rows)}")
                    print(f"DEBUG: form_data keys: {list(form_data.keys()) if isinstance(form_data, dict) else 'Not a dict'}")
                    print(f"DEBUG: rows content: {rows}")
                    form_id = store_exception_form(form_data, rows, username, form_type=form_type, upload_date=upload_date)
                    if not form_id:
                        failed += 1
                        continue
                    # Safely get pass_number for audit log
                    pass_number = 'N/A'
                    if isinstance(form_data, dict):
                        pass_number = form_data.get('pass_number', 'N/A')
                    log_audit(username, 'upload', 'form', form_id, f"Form uploaded: {pass_number}")
                    success += 1
            except Exception as e:
                print(f"Error processing file {file.filename}: {e}")
                print(f"DEBUG: Exception details - form_data type: {type(form_data) if 'form_data' in locals() else 'Not defined'}")
                print(f"DEBUG: Exception details - rows type: {type(rows) if 'rows' in locals() else 'Not defined'}")
                print(f"DEBUG: Exception details - username: {username}")
                print(f"DEBUG: Exception details - form_data content: {form_data if 'form_data' in locals() else 'Not defined'}")
                print(f"DEBUG: Exception details - rows content: {rows if 'rows' in locals() else 'Not defined'}")
                import traceback
                print(f"DEBUG: Full traceback: {traceback.format_exc()}")
                failed += 1
        return jsonify({'message': 'Batch upload complete', 'success': success, 'failed': failed})
    except Exception as e:
        print(f"Error in handle_upload: {e}")
        return jsonify({'error': str(e)}), 500

# @app.route('/api/register', methods=['POST'])
# def register():
#     data = request.json
#     print("Register data received:", data)
#     if add_user(data['username'], data['password']):
#         print(f"User {data['username']} registered successfully")
#         return jsonify({"message": "User registered successfully"}), 201
#     print(f"User {data['username']} registration failed: already exists")
#     return jsonify({"error": "Username already exists"}), 409

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    print("Login data received:", data)
    if check_user(data['username'], data['password']):
        print("Login successful")
        # Dummy token and user info for now
        token = "dummy-token-123"
        user = {
            "id": data['username'],
            "name": data['username'],  # Replace with real name if available
            "bscId": data['username'],
            "role": "user"
        }
        return jsonify({
            "message": "Login successful!",
            "token": token,
            "user": user
        }), 200
    print("Login failed")
    return jsonify({"error": "Invalid credentials"}), 401

@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    print("Registration data received:", data)
    
    # Validate input
    if not data.get('username') or not data.get('password'):
        return jsonify({"error": "Username and password are required"}), 400
    
    if len(data['username']) < 3:
        return jsonify({"error": "Username must be at least 3 characters long"}), 400
    
    if len(data['password']) < 6:
        return jsonify({"error": "Password must be at least 6 characters long"}), 400
    
    # Try to add the user
    if add_user(data['username'], data['password']):
        print("Registration successful")
        return jsonify({
            "message": "Registration successful! You can now login.",
            "user": {
                "id": data['username'],
                "name": data['username'],
                "bscId": data['username'],
                "role": "user"
            }
        }), 201
    else:
        print("Registration failed - username already exists")
        return jsonify({"error": "Username already exists"}), 409

@app.route('/api/stats', methods=['GET'])
def get_stats():
    with sqlite3.connect('forms.db', timeout=10) as conn:
        c = conn.cursor()
        # Sum overtime (convert HH:MM to minutes, then back to HH:MM)
        c.execute('SELECT overtime_hh, overtime_mm FROM exception_form_rows')
        total_minutes = 0
        for hh, mm in c.fetchall():
            try:
                total_minutes += int(hh or 0) * 60 + int(mm or 0)
            except Exception:
                continue
        total_overtime_hh = total_minutes // 60
        total_overtime_mm = total_minutes % 60
        # Count unique TA job numbers (non-empty)
        c.execute('SELECT COUNT(DISTINCT ta_job_no) FROM exception_form_rows WHERE ta_job_no != ""')
        total_job_numbers = c.fetchone()[0]
        conn.commit()
    return jsonify({
        'total_overtime': f"{total_overtime_hh:02d}:{total_overtime_mm:02d}",
        'total_job_numbers': total_job_numbers
    })

@app.route('/api/dashboard', methods=['GET'])
def get_dashboard_data():
    try:
        form_type = request.args.get('form_type')
        username = request.args.get('username')
        extraction_mode_filter = request.args.get('extraction_mode')  # Get extraction mode filter
        
        with sqlite3.connect('forms.db', timeout=10) as conn:
            c = conn.cursor()
            
            # Get forms for statistics calculation with extraction mode filter
            if form_type:
                if extraction_mode_filter == 'pure':
                    c.execute('SELECT id, pass_number, title, employee_name, rdos, actual_ot_date, div, comments, supervisor_name, supervisor_pass_no, oto, oto_amount_saved, entered_in_uts, regular_assignment, report, relief, todays_date, status, username, ocr_lines, form_type, upload_date, file_name, reg, superintendent_authorization_signature, superintendent_authorization_pass, superintendent_authorization_date, entered_into_uts, raw_gemini_json, overtime_hours, report_loc, overtime_location, report_time, relief_time, date_of_overtime, job_number, rc_number, acct_number, reason_rdo, reason_absentee_coverage, reason_no_lunch, reason_early_report, reason_late_clear, reason_save_as_oto, reason_capital_support_go, reason_other, amount, raw_extracted_data, extraction_mode, raw_extracted_data_pure, raw_extracted_data_mapped FROM exception_forms WHERE status = "processed" AND form_type = ? AND (extraction_mode = ? OR extraction_mode = "combined")', (form_type, extraction_mode_filter))
                elif extraction_mode_filter == 'mapped':
                    c.execute('SELECT id, pass_number, title, employee_name, rdos, actual_ot_date, div, comments, supervisor_name, supervisor_pass_no, oto, oto_amount_saved, entered_in_uts, regular_assignment, report, relief, todays_date, status, username, ocr_lines, form_type, upload_date, file_name, reg, superintendent_authorization_signature, superintendent_authorization_pass, superintendent_authorization_date, entered_into_uts, raw_gemini_json, overtime_hours, report_loc, overtime_location, report_time, relief_time, date_of_overtime, job_number, rc_number, acct_number, reason_rdo, reason_absentee_coverage, reason_no_lunch, reason_early_report, reason_late_clear, reason_save_as_oto, reason_capital_support_go, reason_other, amount, raw_extracted_data, extraction_mode, raw_extracted_data_pure, raw_extracted_data_mapped FROM exception_forms WHERE status = "processed" AND form_type = ? AND (extraction_mode = ? OR extraction_mode = "combined" OR extraction_mode IS NULL)', (form_type, extraction_mode_filter))
                else:
                    c.execute('SELECT id, pass_number, title, employee_name, rdos, actual_ot_date, div, comments, supervisor_name, supervisor_pass_no, oto, oto_amount_saved, entered_in_uts, regular_assignment, report, relief, todays_date, status, username, ocr_lines, form_type, upload_date, file_name, reg, superintendent_authorization_signature, superintendent_authorization_pass, superintendent_authorization_date, entered_into_uts, raw_gemini_json, overtime_hours, report_loc, overtime_location, report_time, relief_time, date_of_overtime, job_number, rc_number, acct_number, reason_rdo, reason_absentee_coverage, reason_no_lunch, reason_early_report, reason_late_clear, reason_save_as_oto, reason_capital_support_go, reason_other, amount, raw_extracted_data, extraction_mode, raw_extracted_data_pure, raw_extracted_data_mapped FROM exception_forms WHERE status = "processed" AND form_type = ?', (form_type,))
            else:
                if extraction_mode_filter == 'pure':
                    c.execute('SELECT id, pass_number, title, employee_name, rdos, actual_ot_date, div, comments, supervisor_name, supervisor_pass_no, oto, oto_amount_saved, entered_in_uts, regular_assignment, report, relief, todays_date, status, username, ocr_lines, form_type, upload_date, file_name, reg, superintendent_authorization_signature, superintendent_authorization_pass, superintendent_authorization_date, entered_into_uts, raw_gemini_json, overtime_hours, report_loc, overtime_location, report_time, relief_time, date_of_overtime, job_number, rc_number, acct_number, reason_rdo, reason_absentee_coverage, reason_no_lunch, reason_early_report, reason_late_clear, reason_save_as_oto, reason_capital_support_go, reason_other, amount, raw_extracted_data, extraction_mode, raw_extracted_data_pure, raw_extracted_data_mapped FROM exception_forms WHERE status = "processed" AND (extraction_mode = ? OR extraction_mode = "combined")', (extraction_mode_filter,))
                elif extraction_mode_filter == 'mapped':
                    c.execute('SELECT id, pass_number, title, employee_name, rdos, actual_ot_date, div, comments, supervisor_name, supervisor_pass_no, oto, oto_amount_saved, entered_in_uts, regular_assignment, report, relief, todays_date, status, username, ocr_lines, form_type, upload_date, file_name, reg, superintendent_authorization_signature, superintendent_authorization_pass, superintendent_authorization_date, entered_into_uts, raw_gemini_json, overtime_hours, report_loc, overtime_location, report_time, relief_time, date_of_overtime, job_number, rc_number, acct_number, reason_rdo, reason_absentee_coverage, reason_no_lunch, reason_early_report, reason_late_clear, reason_save_as_oto, reason_capital_support_go, reason_other, amount, raw_extracted_data, extraction_mode, raw_extracted_data_pure, raw_extracted_data_mapped FROM exception_forms WHERE status = "processed" AND (extraction_mode = ? OR extraction_mode = "combined" OR extraction_mode IS NULL)', (extraction_mode_filter,))
                else:
                    c.execute('SELECT id, pass_number, title, employee_name, rdos, actual_ot_date, div, comments, supervisor_name, supervisor_pass_no, oto, oto_amount_saved, entered_in_uts, regular_assignment, report, relief, todays_date, status, username, ocr_lines, form_type, upload_date, file_name, reg, superintendent_authorization_signature, superintendent_authorization_pass, superintendent_authorization_date, entered_into_uts, raw_gemini_json, overtime_hours, report_loc, overtime_location, report_time, relief_time, date_of_overtime, job_number, rc_number, acct_number, reason_rdo, reason_absentee_coverage, reason_no_lunch, reason_early_report, reason_late_clear, reason_save_as_oto, reason_capital_support_go, reason_other, amount, raw_extracted_data, extraction_mode, raw_extracted_data_pure, raw_extracted_data_mapped FROM exception_forms WHERE status = "processed"')
            
            forms = c.fetchall()
            
            # Calculate statistics using the new filtered function
            stats = calculate_dashboard_stats_with_raw_data(forms, form_type, extraction_mode_filter)
            
            # Get forms for the table display with extraction mode filter
            if form_type:
                if extraction_mode_filter == 'pure':
                    c.execute('SELECT id, pass_number, title, employee_name, actual_ot_date, div, comments, status, form_type, upload_date FROM exception_forms WHERE status = "processed" AND form_type = ? AND (extraction_mode = ? OR extraction_mode = "combined")', (form_type, extraction_mode_filter))
                elif extraction_mode_filter == 'mapped':
                    c.execute('SELECT id, pass_number, title, employee_name, actual_ot_date, div, comments, status, form_type, upload_date FROM exception_forms WHERE status = "processed" AND form_type = ? AND (extraction_mode = ? OR extraction_mode = "combined" OR extraction_mode IS NULL)', (form_type, extraction_mode_filter))
                else:
                    c.execute('SELECT id, pass_number, title, employee_name, actual_ot_date, div, comments, status, form_type, upload_date FROM exception_forms WHERE status = "processed" AND form_type = ?', (form_type,))
            else:
                if extraction_mode_filter == 'pure':
                    c.execute('SELECT id, pass_number, title, employee_name, actual_ot_date, div, comments, status, form_type, upload_date FROM exception_forms WHERE status = "processed" AND (extraction_mode = ? OR extraction_mode = "combined")', (extraction_mode_filter,))
                elif extraction_mode_filter == 'mapped':
                    c.execute('SELECT id, pass_number, title, employee_name, actual_ot_date, div, comments, status, form_type, upload_date FROM exception_forms WHERE status = "processed" AND (extraction_mode = ? OR extraction_mode = "combined" OR extraction_mode IS NULL)', (extraction_mode_filter,))
                else:
                    c.execute('SELECT id, pass_number, title, employee_name, actual_ot_date, div, comments, status, form_type, upload_date FROM exception_forms WHERE status = "processed"')
            
            forms_table = []
            for row in c.fetchall():
                # Get raw_extracted_data and extraction_mode for this form
                form_id = row[0]
                c2 = conn.cursor()
                c2.execute('SELECT raw_extracted_data, raw_extracted_data_pure, raw_extracted_data_mapped, extraction_mode FROM exception_forms WHERE id = ?', (form_id,))
                raw_data_row = c2.fetchone()
                raw_json = None
                form_extraction_mode = None
                if raw_data_row:
                    raw_json = None
                    form_extraction_mode = raw_data_row[3]
                    
                    # Select the appropriate raw data based on extraction mode
                    if extraction_mode_filter == 'pure' and form_extraction_mode == 'combined':
                        raw_data = raw_data_row[1]  # raw_extracted_data_pure
                    elif extraction_mode_filter == 'mapped' and form_extraction_mode == 'combined':
                        raw_data = raw_data_row[2]  # raw_extracted_data_mapped
                    else:
                        raw_data = raw_data_row[0]  # raw_extracted_data (legacy)
                    
                    if raw_data:
                        try:
                            raw_json = json.loads(raw_data)
                        except Exception:
                            raw_json = None
                
                def get_field(field, fallback_idx):
                    # Handle combined extraction mode
                    if form_extraction_mode == 'combined':
                        if extraction_mode_filter == 'pure':
                            # Use pure extraction data
                            if raw_json and field in raw_json:
                                print(f"[DEBUG] Using RAW value for {field} (form {form_id}) in pure mode: {raw_json[field]}")
                                return raw_json[field]
                        elif extraction_mode_filter == 'mapped':
                            # Use mapped extraction data
                            if raw_json and field in raw_json:
                                print(f"[DEBUG] Using RAW value for {field} (form {form_id}) in mapped mode: {raw_json[field]}")
                                return raw_json[field]
                    
                    # Handle legacy extraction modes
                    elif (extraction_mode_filter == 'pure' and 
                          form_extraction_mode == 'pure' and 
                          raw_json and field in raw_json):
                        print(f"[DEBUG] Using RAW value for {field} (form {form_id}): {raw_json[field]}")
                        return raw_json[field]
                    elif (extraction_mode_filter == 'mapped' and 
                          form_extraction_mode == 'mapped' and 
                          raw_json and field in raw_json):
                        print(f"[DEBUG] Using RAW value for {field} (form {form_id}) in mapped mode: {raw_json[field]}")
                        return raw_json[field]
                    
                    # Fallback to database fields
                    print(f"[DEBUG] Using FALLBACK value for {field} (form {form_id}): {row[fallback_idx]}")
                    return row[fallback_idx]
                
                forms_table.append({
                    "id": row[0],
                    "pass_number": get_field("pass_number", 1),
                    "title": get_field("title", 2),
                    "employee_name": get_field("employee_name", 3),
                    "actual_ot_date": get_field("actual_ot_date", 4),
                    "div": get_field("div", 5),
                    "comments": get_field("comments", 6),
                    "status": row[7],
                    "form_type": row[8] if len(row) > 8 else '',
                    "upload_date": row[9] if len(row) > 9 else ''
                })
            # DEBUG: Print the final forms_table
            print("[DEBUG] Final forms_table for dashboard:")
            for f in forms_table:
                print(f)
            
            conn.commit()
        
        result = {
            "total_forms": stats["total_forms"],
            "total_overtime": stats["total_overtime"],
            "unique_job_numbers": stats["unique_job_numbers"],
            "most_relevant_position": stats["most_relevant_position"],
            "most_relevant_location": stats["most_relevant_location"],
            "most_common_reason": stats.get("most_common_reason"),
            "forms": forms_table
        }
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error in dashboard: {e}")
        return jsonify({"error": str(e)}), 500

def calculate_dashboard_stats_with_raw_data(forms, form_type=None, extraction_mode_filter=None):
    """
    Calculate dashboard statistics using ONLY pure extraction data.
    All stats are derived from raw_extracted_data JSON, ignoring mapped fields entirely.
    """
    import json
    from collections import Counter
    
    def safe_int(val):
        try:
            return int(val)
        except Exception:
            return 0
    
    total_forms = len(forms)
    total_minutes = 0
    job_numbers = []
    positions = []
    locations = []
    reason_counts = {field: 0 for field in [
        'reason_rdo', 'reason_absentee_coverage', 'reason_no_lunch', 
        'reason_early_report', 'reason_late_clear', 'reason_save_as_oto', 
        'reason_capital_support_go', 'reason_other'
    ]}
    
    # Define the exact column order that matches our SELECT statement
    columns = ['id', 'pass_number', 'title', 'employee_name', 'rdos', 'actual_ot_date', 'div', 'comments', 'supervisor_name', 'supervisor_pass_no', 'oto', 'oto_amount_saved', 'entered_in_uts', 'regular_assignment', 'report', 'relief', 'todays_date', 'status', 'username', 'ocr_lines', 'form_type', 'upload_date', 'file_name', 'reg', 'superintendent_authorization_signature', 'superintendent_authorization_pass', 'superintendent_authorization_date', 'entered_into_uts', 'raw_gemini_json', 'overtime_hours', 'report_loc', 'overtime_location', 'report_time', 'relief_time', 'date_of_overtime', 'job_number', 'rc_number', 'acct_number', 'reason_rdo', 'reason_absentee_coverage', 'reason_no_lunch', 'reason_early_report', 'reason_late_clear', 'reason_save_as_oto', 'reason_capital_support_go', 'reason_other', 'amount', 'raw_extracted_data', 'extraction_mode', 'raw_extracted_data_pure', 'raw_extracted_data_mapped']
    
    for form in forms:
        # Convert tuple to dictionary using column names
        if isinstance(form, tuple):
            form_dict = dict(zip(columns, form))
        else:
            form_dict = dict(form)
        
        print(f"\n--- Debug: Processing form ID {form_dict.get('id')} ---")
        print(f"Extraction mode: {form_dict.get('extraction_mode')}")
        print(f"Raw extracted data: {form_dict.get('raw_extracted_data')}")
        
        extraction_mode = form_dict.get('extraction_mode')
        current_form_type = form_dict.get('form_type')
        
        # Select the appropriate raw data based on extraction mode
        if extraction_mode == 'combined':
            if extraction_mode_filter == 'pure':
                raw_data = form_dict.get('raw_extracted_data_pure')
            elif extraction_mode_filter == 'mapped':
                raw_data = form_dict.get('raw_extracted_data_mapped')
            else:
                # Default to mapped data for combined forms
                raw_data = form_dict.get('raw_extracted_data_mapped')
        else:
            raw_data = form_dict.get('raw_extracted_data')
        
        # Filter by extraction mode if specified
        if extraction_mode_filter:
            if extraction_mode_filter == 'pure':
                # For pure mode, show forms that are 'pure' OR 'combined'
                if extraction_mode not in ['pure', 'combined']:
                    print(f"Skipping form {form_dict.get('id')} - extraction mode mismatch (form: {extraction_mode}, filter: {extraction_mode_filter})")
                    continue
            elif extraction_mode_filter == 'mapped':
                # For mapped mode, show forms that are 'mapped', 'combined', or NULL/empty (legacy)
                if extraction_mode not in [None, '', 'mapped', 'combined']:
                    print(f"Skipping form {form_dict.get('id')} - extraction mode mismatch (form: {extraction_mode}, filter: {extraction_mode_filter})")
                    continue
        
        # Use the current form's type for processing this specific form
        current_form_type_for_processing = current_form_type
        
        # For pure extraction mode, ONLY process forms that have raw_extracted_data
        # For mapped extraction mode, process all forms (they may or may not have raw_extracted_data)
        if extraction_mode_filter == 'pure':
            if not raw_data:
                print(f"Skipping form {form_dict.get('id')} - no raw_extracted_data in pure mode")
                continue
        elif extraction_mode_filter == 'mapped':
            # For mapped mode, we can process forms even without raw_extracted_data
            pass
        
        # Process the form data
        if raw_data:
            try:
                raw_json = json.loads(raw_data)
                
                # --- Flexible Overtime Extraction ---
                if current_form_type_for_processing == 'supervisor':
                    overtime_hours = get_flexible_field(raw_json, [
                        'overtime_hours', 'overtime', 'hours', 'ot_hours', 'ot', 'total_overtime'
                    ])
                    if overtime_hours and overtime_hours != 'N/A':
                        try:
                            if '+' in str(overtime_hours):
                                parts = str(overtime_hours).split('+')
                                for part in parts:
                                    part = part.strip()
                                    if ':' in part:
                                        hours, minutes = part.split(':')
                                        total_minutes += int(hours) * 60 + int(minutes)
                                    else:
                                        total_minutes += int(part) * 60
                            elif ':' in str(overtime_hours):
                                hours, minutes = str(overtime_hours).split(':')
                                total_minutes += int(hours) * 60 + int(minutes)
                            else:
                                total_minutes += int(overtime_hours) * 60
                        except:
                            pass
                    
                    # --- Flexible Job Number Extraction (Supervisor) ---
                    job_num = get_flexible_field(raw_json, [
                        'job', 'job_number', 'job no', 'job_no', 'job#', 'job number', 'ta_job_no', 'ta job no', 'jobnum', 'jobnumber', 'job id', 'jobid'
                    ])
                    if job_num and job_num != 'N/A':
                        job_numbers.append(str(job_num))
                elif current_form_type_for_processing == 'hourly':
                    # Look for overtime in rows or direct fields
                    rows = raw_json.get('rows', [])
                    for row in rows:
                        if isinstance(row, dict):
                            # --- Flexible Overtime Extraction (Hourly Row) ---
                            hh = get_flexible_field(row, ['overtime_hh', 'ot_hh', 'hh'])
                            mm = get_flexible_field(row, ['overtime_mm', 'ot_mm', 'mm'])
                            print('Hourly row:', row)
                            print('Extracted overtime:', hh, mm)
                            try:
                                total_minutes += safe_int(hh) * 60 + safe_int(mm)
                            except:
                                pass
                            # --- Flexible Job Number Extraction (Hourly Row) ---
                            ta_job = get_flexible_field(row, [
                                'ta_job_no', 'job_no', 'jobnumber', 'jobnum'
                            ])
                            if ta_job and ta_job != 'N/A':
                                job_numbers.append(str(ta_job))
                            # --- Flexible Location Extraction (Hourly Row) ---
                            line_loc = get_flexible_field(row, [
                                'Line/Location *', 'Line/Location', 'line_location', 'location', 'line', 'loc', 'line_loc'
                            ])
                            if line_loc and line_loc != 'N/A':
                                locations.append(line_loc)
                    # Also check for overtime, job number, and location at the top level (in case some forms store them there)
                    hh_top = get_flexible_field(raw_json, ['overtime_hh', 'ot_hh', 'hh'])
                    mm_top = get_flexible_field(raw_json, ['overtime_mm', 'ot_mm', 'mm'])
                    try:
                        total_minutes += safe_int(hh_top) * 60 + safe_int(mm_top)
                    except:
                        pass
                    # Also check for overtime_hours at the top level (e.g., "00:30")
                    overtime_hours_top = get_flexible_field(raw_json, [
                        'overtime_hours', 'overtime', 'hours', 'ot_hours', 'ot', 'total_overtime'
                    ])
                    if overtime_hours_top and overtime_hours_top != 'N/A':
                        try:
                            if '+' in str(overtime_hours_top):
                                parts = str(overtime_hours_top).split('+')
                                for part in parts:
                                    part = part.strip()
                                    if ':' in part:
                                        hours, minutes = part.split(':')
                                        total_minutes += int(hours) * 60 + int(minutes)
                                    else:
                                        total_minutes += int(part) * 60
                            elif ':' in str(overtime_hours_top):
                                hours, minutes = str(overtime_hours_top).split(':')
                                total_minutes += int(hours) * 60 + int(minutes)
                            else:
                                total_minutes += int(overtime_hours_top) * 60
                        except:
                            pass
                    # --- NEW: Also check for job number and location at the top level ---
                    ta_job_top = get_flexible_field(raw_json, [
                        'ta_job_no', 'job_no', 'jobnumber', 'jobnum', 'job number', 'TA Job No', 'TA Job Number'
                    ])
                    if ta_job_top and ta_job_top != 'N/A':
                        job_numbers.append(str(ta_job_top))
                    line_loc_top = get_flexible_field(raw_json, [
                        'Line/Location *', 'Line/Location', 'line_location', 'location', 'line', 'loc', 'line_loc', 'report_station'
                    ])
                    if line_loc_top and line_loc_top != 'N/A':
                        locations.append(line_loc_top)
                
                # --- Flexible Position/Title Extraction ---
                if current_form_type_for_processing == 'supervisor':
                    positions.append('Supervisor')
                else:
                    title = get_flexible_field(raw_json, [
                        'title', 'position', 'job_title', 'role'
                    ])
                    if title and title != 'N/A':
                        positions.append(title)
                
                # --- Flexible Location Extraction ---
                if current_form_type_for_processing == 'supervisor':
                    # Check for pure extraction field names first (same as hourly)
                    line_loc = get_flexible_field(raw_json, [
                        'Line/Location *', 'Line/Location', 'line_location', 'location', 'line', 'loc', 'line_loc'
                    ])
                    if line_loc and line_loc != 'N/A':
                        locations.append(line_loc)
                    else:
                        # Fallback to supervisor-specific field names
                        report_loc = get_flexible_field(raw_json, [
                            'report_loc', 'report_location', 'location', 'reportloc', 'report', 'loc'
                        ])
                        overtime_loc = get_flexible_field(raw_json, [
                            'overtime_location', 'location', 'overtimelocation', 'ot_location', 'otloc'
                        ])
                        # Only add unique locations per form to avoid double counting
                        form_locations = set()
                        if report_loc and report_loc != 'N/A':
                            form_locations.add(report_loc)
                        if overtime_loc and overtime_loc != 'N/A':
                            form_locations.add(overtime_loc)
                        # Add unique locations from this form
                        for loc in form_locations:
                            locations.append(loc)
                elif current_form_type_for_processing == 'hourly':
                    rows = raw_json.get('rows', [])
                    for row in rows:
                        if isinstance(row, dict):
                            line_loc = get_flexible_field(row, [
                                'Line/Location *', 'Line/Location', 'line_location', 'location', 'line', 'loc', 'line_loc'
                            ])
                            if line_loc and line_loc != 'N/A':
                                locations.append(line_loc)
                
                # --- Flexible Reason Extraction (Supervisor) ---
                if current_form_type_for_processing == 'supervisor':
                    reasons = get_flexible_field(raw_json, [
                        'reason_for_overtime', 'reason', 'overtime_reason', 'reasonovertime', 'reasonforot'
                    ])
                    if isinstance(reasons, str):
                        reasons = [reasons]
                    if isinstance(reasons, list):
                        for reason in reasons:
                            reason_lower = str(reason).lower()
                            if 'rdo' in reason_lower:
                                reason_counts['reason_rdo'] += 1
                            elif 'absentee' in reason_lower or 'coverage' in reason_lower:
                                reason_counts['reason_absentee_coverage'] += 1
                            elif 'lunch' in reason_lower:
                                reason_counts['reason_no_lunch'] += 1
                            elif 'early' in reason_lower and 'report' in reason_lower:
                                reason_counts['reason_early_report'] += 1
                            elif 'late' in reason_lower and 'clear' in reason_lower:
                                reason_counts['reason_late_clear'] += 1
                            elif 'oto' in reason_lower:
                                reason_counts['reason_save_as_oto'] += 1
                            elif 'capital' in reason_lower or 'support' in reason_lower:
                                reason_counts['reason_capital_support_go'] += 1
                            else:
                                reason_counts['reason_other'] += 1
            except json.JSONDecodeError:
                # Skip forms with invalid JSON
                print(f"Error parsing JSON for form {form_dict.get('id')}")
                continue
        elif extraction_mode_filter == 'mapped':
            # For mapped extraction mode, use the mapped fields from the database
            print(f"Processing form {form_dict.get('id')} in mapped mode using database fields")
            
            # --- Flexible Overtime Extraction from mapped fields ---
            if current_form_type_for_processing == 'supervisor':
                overtime_hours = form_dict.get('overtime_hours')
                if overtime_hours and overtime_hours != 'N/A':
                    try:
                        if '+' in str(overtime_hours):
                            parts = str(overtime_hours).split('+')
                            for part in parts:
                                part = part.strip()
                                if ':' in part:
                                    hours, minutes = part.split(':')
                                    total_minutes += int(hours) * 60 + int(minutes)
                                else:
                                    total_minutes += int(part) * 60
                        elif ':' in str(overtime_hours):
                            hours, minutes = str(overtime_hours).split(':')
                            total_minutes += int(hours) * 60 + int(minutes)
                        else:
                            total_minutes += int(overtime_hours) * 60
                    except:
                        pass
                
                # --- Job Number from mapped fields ---
                job_num = form_dict.get('job_number')
                if job_num and job_num != 'N/A':
                    job_numbers.append(str(job_num))
                
                # --- Location from mapped fields ---
                report_loc = form_dict.get('report_loc')
                overtime_loc = form_dict.get('overtime_location')
                if report_loc and report_loc != 'N/A':
                    locations.append(report_loc)
                if overtime_loc and overtime_loc != 'N/A':
                    locations.append(overtime_loc)
                
                # --- Position/Title from mapped fields ---
                title = form_dict.get('title')
                if title and title != 'N/A':
                    positions.append(title)
                else:
                    positions.append('Supervisor')
                
                # --- Reason from mapped fields ---
                # Check all reason fields and count them
                for reason_field in ['reason_rdo', 'reason_absentee_coverage', 'reason_no_lunch', 
                                   'reason_early_report', 'reason_late_clear', 'reason_save_as_oto', 
                                   'reason_capital_support_go', 'reason_other']:
                    if form_dict.get(reason_field):
                        reason_counts[reason_field] += 1
    
    # Calculate final statistics
    total_overtime_hh = total_minutes // 60
    total_overtime_mm = total_minutes % 60
    
    unique_job_numbers = set(job_numbers)
    
    most_position, most_position_count = ('N/A', 0)
    if positions:
        most_position, most_position_count = Counter(positions).most_common(1)[0]
    
    most_location, most_location_count = ('N/A', 0)
    if locations:
        most_location, most_location_count = Counter(locations).most_common(1)[0]
    
    most_common_reason = None
    if any(reason_counts.values()):
        most_common_reason_field = max(reason_counts, key=lambda k: reason_counts[k])
        reason_label_map = {
            'reason_rdo': 'RDO',
            'reason_absentee_coverage': 'Absentee Coverage',
            'reason_no_lunch': 'No Lunch',
            'reason_early_report': 'Early Report',
            'reason_late_clear': 'Late Clear',
            'reason_save_as_oto': 'Save as OTO',
            'reason_capital_support_go': 'Capital Support / GO',
            'reason_other': 'Other'
        }
        most_common_reason = {
            'reason': reason_label_map.get(most_common_reason_field, most_common_reason_field),
            'count': reason_counts[most_common_reason_field]
        }
    else:
        most_common_reason = {'reason': 'N/A', 'count': 0}
    
    return {
        "total_forms": total_forms,
        "total_overtime": f"{total_overtime_hh}h {total_overtime_mm}m",
        "unique_job_numbers": len(unique_job_numbers),
        "most_relevant_position": {"position": most_position, "count": most_position_count},
        "most_relevant_location": {"location": most_location, "count": most_location_count},
        "most_common_reason": most_common_reason
    }

@app.route('/api/form/<int:form_id>', methods=['GET'])
def get_form_details(form_id):
    import sqlite3
    import json
    extraction_mode = request.args.get('extraction_mode', 'mapped')
    
    with sqlite3.connect('forms.db', timeout=10) as conn:
        c = conn.cursor()
        # Get form header - don't filter by extraction mode for existing forms
        # This allows access to legacy forms that only exist in one extraction mode
        c.execute('SELECT * FROM exception_forms WHERE id = ?', (form_id,))
        
        form_row = c.fetchone()
        if not form_row:
            return jsonify({'error': 'Form not found'}), 404
        
        columns = [desc[0] for desc in c.description]
        form_data = dict(zip(columns, form_row))
        
        # Get the raw extracted data based on extraction mode
        raw_json = None
        form_extraction_mode = form_data.get('extraction_mode')
        
        if form_extraction_mode == 'combined':
            if extraction_mode == 'pure':
                raw_data = form_data.get('raw_extracted_data_pure')
            elif extraction_mode == 'mapped':
                raw_data = form_data.get('raw_extracted_data_mapped')
            else:
                raw_data = form_data.get('raw_extracted_data_mapped')  # Default to mapped
        else:
            raw_data = form_data.get('raw_extracted_data')
        
        if raw_data:
            try:
                raw_json = json.loads(raw_data)
            except json.JSONDecodeError:
                print(f"Error parsing JSON for form {form_id}: {raw_data}")
        
        # Function to get field value with proper extraction mode handling
        def get_field(field, fallback_value):
            # Handle combined extraction mode
            if form_extraction_mode == 'combined':
                if extraction_mode == 'pure':
                    # Use pure extraction data
                    if raw_json and field in raw_json:
                        return raw_json[field]
                elif extraction_mode == 'mapped':
                    # Use mapped extraction data
                    if raw_json and field in raw_json:
                        return raw_json[field]
            
            # Handle legacy extraction modes
            elif (extraction_mode == 'pure' and 
                  form_extraction_mode == 'pure' and 
                  raw_json and field in raw_json):
                return raw_json[field]
            elif (extraction_mode == 'mapped' and 
                  form_extraction_mode == 'mapped' and 
                  raw_json and field in raw_json):
                return raw_json[field]
            
            # Fallback to database fields
            return fallback_value
        
        # Update form_data with proper field values based on extraction mode
        if extraction_mode == 'pure' and raw_json:
            # For pure extraction, use raw Gemini data directly
            if form_data.get('form_type') == 'supervisor':
                # Use raw supervisor field names from Gemini
                form_data['pass_number'] = raw_json.get('PASS', form_data.get('pass_number', ''))
                form_data['title'] = raw_json.get('TITLE', form_data.get('title', ''))
                form_data['employee_name'] = raw_json.get('EMPLOYEE NAME', form_data.get('employee_name', ''))
                form_data['actual_ot_date'] = raw_json.get('DATE OF OVERTIME', form_data.get('actual_ot_date', ''))
                form_data['div'] = raw_json.get('DIV', form_data.get('div', ''))
                form_data['comments'] = raw_json.get('COMMENTS', form_data.get('comments', ''))
            else:
                # For hourly forms, use standard field names
                form_data['pass_number'] = raw_json.get('pass_number', form_data.get('pass_number', ''))
                form_data['title'] = raw_json.get('title', form_data.get('title', ''))
                form_data['employee_name'] = raw_json.get('employee_name', form_data.get('employee_name', ''))
                form_data['actual_ot_date'] = raw_json.get('actual_ot_date', form_data.get('actual_ot_date', ''))
                form_data['div'] = raw_json.get('div', form_data.get('div', ''))
                form_data['comments'] = raw_json.get('comments', form_data.get('comments', ''))
        else:
            # For mapped extraction, use the get_field function
            form_data['pass_number'] = get_field('pass_number', form_data.get('pass_number', ''))
            form_data['title'] = get_field('title', form_data.get('title', ''))
            form_data['employee_name'] = get_field('employee_name', form_data.get('employee_name', ''))
            form_data['actual_ot_date'] = get_field('actual_ot_date', form_data.get('actual_ot_date', ''))
            form_data['div'] = get_field('div', form_data.get('div', ''))
            form_data['comments'] = get_field('comments', form_data.get('comments', ''))
        
        # Get all rows for this form
        c.execute('SELECT * FROM exception_form_rows WHERE form_id = ?', (form_id,))
        rows = c.fetchall()
        row_columns = [desc[0] for desc in c.description]
        form_rows = [dict(zip(row_columns, row)) for row in rows]
        conn.commit()
    return jsonify({'form': form_data, 'rows': form_rows})

@app.route('/api/form/<int:form_id>', methods=['PUT'])
def update_form(form_id):
    import sqlite3
    data = request.json
    form = data.get('form', {})
    rows = data.get('rows', [])
    username = data.get('username', 'unknown')  # For audit logging
    with sqlite3.connect('forms.db', timeout=10) as conn:
        c = conn.cursor()
        # Update form header
        c.execute('''
            UPDATE exception_forms SET
                pass_number = ?,
                title = ?,
                employee_name = ?,
                rdos = ?,
                actual_ot_date = ?,
                div = ?,
                comments = ?,
                supervisor_name = ?,
                supervisor_pass_no = ?,
                oto = ?,
                oto_amount_saved = ?,
                entered_in_uts = ?,
                regular_assignment = ?,
                report = ?,
                relief = ?,
                todays_date = ?,
                status = ?,
                reg = ?,
                superintendent_authorization_signature = ?,
                superintendent_authorization_pass = ?,
                superintendent_authorization_date = ?,
                entered_into_uts = ?,
                overtime_hours = ?,
                report_loc = ?,
                overtime_location = ?,
                report_time = ?,
                relief_time = ?,
                date_of_overtime = ?,
                job_number = ?,
                rc_number = ?,
                acct_number = ?,
                amount = ?,
                reason_rdo = ?,
                reason_absentee_coverage = ?,
                reason_no_lunch = ?,
                reason_early_report = ?,
                reason_late_clear = ?,
                reason_save_as_oto = ?,
                reason_capital_support_go = ?,
                reason_other = ?
            WHERE id = ?
        ''', (
            form.get('pass_number', ''),
            form.get('title', ''),
            form.get('employee_name', ''),
            form.get('rdos', ''),
            form.get('actual_ot_date', ''),
            form.get('div', ''),
            form.get('comments', ''),
            form.get('supervisor_name', ''),
            form.get('supervisor_pass_no', ''),
            form.get('oto', ''),
            form.get('oto_amount_saved', ''),
            form.get('entered_in_uts', ''),
            form.get('regular_assignment', ''),
            form.get('report', ''),
            form.get('relief', ''),
            form.get('todays_date', ''),
            form.get('status', ''),
            form.get('reg', ''),
            form.get('superintendent_authorization_signature', ''),
            form.get('superintendent_authorization_pass', ''),
            form.get('superintendent_authorization_date', ''),
            form.get('entered_into_uts', ''),
            form.get('overtime_hours', ''),
            form.get('report_loc', ''),
            form.get('overtime_location', ''),
            form.get('report_time', ''),
            form.get('relief_time', ''),
            form.get('date_of_overtime', ''),
            form.get('job_number', ''),
            form.get('rc_number', ''),
            form.get('acct_number', ''),
            form.get('amount', ''),
            form.get('reason_rdo', False),
            form.get('reason_absentee_coverage', False),
            form.get('reason_no_lunch', False),
            form.get('reason_early_report', False),
            form.get('reason_late_clear', False),
            form.get('reason_save_as_oto', False),
            form.get('reason_capital_support_go', False),
            form.get('reason_other', False),
            form_id
        ))
        # PATCH: Update raw_gemini_json if present
        if 'raw_gemini_json' in form:
            print(f"Updating raw_gemini_json for form {form_id}: {form['raw_gemini_json'][:100]}...")
            c.execute('UPDATE exception_forms SET raw_gemini_json = ? WHERE id = ?', (form['raw_gemini_json'], form_id))
        # PATCH: Update raw_extracted_data if present
        if 'raw_extracted_data' in form:
            print(f"Updating raw_extracted_data for form {form_id}: {form['raw_extracted_data'][:100]}...")
            c.execute('UPDATE exception_forms SET raw_extracted_data = ? WHERE id = ?', (form['raw_extracted_data'], form_id))
        # Delete old rows
        c.execute('DELETE FROM exception_form_rows WHERE form_id = ?', (form_id,))
        # Insert new rows
        for row in rows:
            c.execute('''
                INSERT INTO exception_form_rows (
                    form_id, code, code_description, line_location, run_no,
                    exception_time_from_hh, exception_time_from_mm,
                    exception_time_to_hh, exception_time_to_mm,
                    overtime_hh, overtime_mm, bonus_hh, bonus_mm,
                    nite_diff_hh, nite_diff_mm, ta_job_no
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                form_id,
                row.get('code', ''),
                row.get('code_description', ''),
                row.get('line_location', ''),
                row.get('run_no', ''),
                row.get('exception_time_from_hh', ''),
                row.get('exception_time_from_mm', ''),
                row.get('exception_time_to_hh', ''),
                row.get('exception_time_to_mm', ''),
                row.get('overtime_hh', ''),
                row.get('overtime_mm', ''),
                row.get('bonus_hh', ''),
                row.get('bonus_mm', ''),
                row.get('nite_diff_hh', ''),
                row.get('nite_diff_mm', ''),
                row.get('ta_job_no', '')
            ))
        conn.commit()
        # Audit log
        c.execute('''
            INSERT INTO audit_trail (username, action, target_type, target_id, details)
            VALUES (?, ?, ?, ?, ?)
        ''', (username, 'edit', 'form', form_id, 'Form and rows edited via API'))
        conn.commit()
    return jsonify({'message': 'Form updated successfully.'})

@app.route('/api/form/<int:form_id>', methods=['DELETE'])
def delete_form(form_id):
    try:
        with sqlite3.connect('forms.db', timeout=10) as conn:
            c = conn.cursor()
            c.execute('DELETE FROM exception_form_rows WHERE form_id = ?', (form_id,))
            c.execute('DELETE FROM exception_forms WHERE id = ?', (form_id,))
            from db import log_audit
            log_audit('system', 'delete', 'form', form_id, "Form deleted via API", conn=conn)
        return jsonify({'message': 'Form deleted successfully.'})
    except Exception as e:
        print(f"Error deleting form {form_id}: {e}")
        return jsonify({'error': str(e)}), 500

def init_audit_db():
    with sqlite3.connect('forms.db', timeout=10) as conn:
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS audit_trail (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT,
                action TEXT,
                target_type TEXT,
                target_id INTEGER,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                details TEXT
            )
        ''')
        conn.commit()

def log_audit(username, action, target_type, target_id, details="", conn=None):
    close_conn = False
    if conn is None:
        conn = sqlite3.connect('forms.db', timeout=10)
        close_conn = True
    c = conn.cursor()
    c.execute('''
        INSERT INTO audit_trail (username, action, target_type, target_id, details)
        VALUES (?, ?, ?, ?, ?)
    ''', (username, action, target_type, target_id, details))
    if close_conn:
        conn.commit()
        conn.close()

@app.route('/api/audit-trail', methods=['GET'])
def get_audit_trail():
    with sqlite3.connect('forms.db', timeout=10) as conn:
        c = conn.cursor()
        c.execute('SELECT id, username, action, target_type, target_id, timestamp, details FROM audit_trail ORDER BY timestamp DESC')
        logs = [
            {
                "id": row[0],
                "user": row[1],
                "action": row[2],
                "target": f"{row[3]}:{row[4]}",
                "timestamp": row[5],
                "details": row[6]
            }
            for row in c.fetchall()
        ]
        conn.commit()
    return jsonify({"logs": logs})

@app.route('/api/extraction-mode', methods=['GET', 'POST'])
def extraction_mode():
    """Get or set the extraction mode (pure vs mapped)"""
    global PURE_GEMINI_EXTRACTION
    
    if request.method == 'POST':
        data = request.get_json()
        if data and 'mode' in data:
            PURE_GEMINI_EXTRACTION = data['mode'] == 'pure'
            return jsonify({
                'mode': 'pure' if PURE_GEMINI_EXTRACTION else 'mapped',
                'message': f'Extraction mode set to {"pure" if PURE_GEMINI_EXTRACTION else "mapped"}'
            })
        else:
            return jsonify({'error': 'Mode not specified'}), 400
    
    return jsonify({
        'mode': 'pure' if PURE_GEMINI_EXTRACTION else 'mapped',
        'description': {
            'pure': 'Extract all fields from Gemini without mapping to predefined fields',
            'mapped': 'Map Gemini output to predefined database fields'
        }
    })

@app.route('/api/forms/export', methods=['GET'])
def export_forms():
    import sqlite3
    import json
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Get query parameters
    form_type = request.args.get('form_type')
    extraction_mode = request.args.get('extraction_mode')
    
    with sqlite3.connect('forms.db', timeout=10) as conn:
        c = conn.cursor()
        
        # Build query with filters
        query = "SELECT * FROM exception_forms WHERE 1=1"
        params = []
        
        if form_type:
            query += " AND form_type = ?"
            params.append(form_type)
        
        if extraction_mode:
            # Handle combined extraction mode - when user wants 'mapped', also include 'combined' forms
            if extraction_mode == 'mapped':
                query += " AND (extraction_mode = ? OR extraction_mode = 'combined')"
                params.append(extraction_mode)
            elif extraction_mode == 'pure':
                query += " AND (extraction_mode = ? OR extraction_mode = 'combined')"
                params.append(extraction_mode)
            else:
                query += " AND extraction_mode = ?"
                params.append(extraction_mode)
        
        c.execute(query, params)
        rows = c.fetchall()
        
        # Get column names
        columns = [desc[0] for desc in c.description]
        
        # Enhanced headers focusing on overtime data
        header_map = {
            'id': 'Form ID',
            'pass_number': 'Pass Number',
            'title': 'Title',
            'employee_name': 'Employee Name',
            'rdos': 'RDOS',
            'actual_ot_date': 'Actual OT Date',
            'div': 'DIV',
            'comments': 'Comments',
            'supervisor_name': 'Supervisor Name',
            'supervisor_pass_no': 'Supervisor Pass No.',
            'oto': 'OTO',
            'oto_amount_saved': 'OTO Amount Saved',
            'entered_in_uts': 'Entered in UTS',
            'regular_assignment': 'Regular Assignment',
            'report': 'Report',
            'relief': 'Relief',
            'todays_date': "Today's Date",
            'form_type': 'Form Type',
            'extraction_mode': 'Extraction Mode',
            'raw_extracted_data': 'Raw Gemini Data (JSON)',
            'raw_gemini_json': 'Raw Gemini JSON',
            'upload_date': 'Upload Date'
        }
        
        # Create enhanced headers with overtime focus
        enhanced_headers = []
        enhanced_columns = []
        
        # First: Essential overtime fields
        overtime_fields = [
            'id', 'pass_number', 'employee_name', 'actual_ot_date', 'form_type',
            'overtime_hours', 'report_loc', 'overtime_location', 'report_time', 'relief_time',
            'date_of_overtime', 'job_number', 'rc_number', 'acct_number', 'amount'
        ]
        
        # Add overtime fields first (if they exist in the database)
        for field in overtime_fields:
            if field in columns:
                enhanced_columns.append(field)
                enhanced_headers.append(header_map.get(field, field.replace('_', ' ').title()))
        
        # Then: Other important fields
        other_fields = [
            'title', 'rdos', 'div', 'comments', 'supervisor_name', 'supervisor_pass_no',
            'oto', 'oto_amount_saved', 'entered_in_uts', 'regular_assignment', 'report', 'relief',
            'todays_date', 'extraction_mode', 'upload_date'
        ]
        
        for field in other_fields:
            if field in columns and field not in enhanced_columns:
                enhanced_columns.append(field)
                enhanced_headers.append(header_map.get(field, field.replace('_', ' ').title()))
        
        # Finally: Raw data fields for debugging
        raw_fields = ['raw_extracted_data', 'raw_gemini_json']
        for field in raw_fields:
            if field in columns:
                enhanced_columns.append(field)
                enhanced_headers.append(header_map.get(field, field.replace('_', ' ').title()))
        
        conn.commit()
    
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Exception Forms"
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add headers with styling
    for col_num, header in enumerate(enhanced_headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data rows with overtime details
    current_row = 2
    for form_row in rows:
            # Create a mapping of column names to values
            form_data = dict(zip(columns, form_row))
            
            # Add main form data
            for col_num, col_name in enumerate(enhanced_columns, 1):
                item = form_data.get(col_name)
                if item is None:
                    cell_value = ''
                elif col_name in ['raw_extracted_data', 'raw_gemini_json'] and item:
                    # Format JSON data for Excel readability
                    try:
                        parsed_json = json.loads(item)
                        formatted_json = json.dumps(parsed_json, indent=2)
                        cell_value = formatted_json
                    except:
                        cell_value = str(item)
                else:
                    cell_value = str(item) if item is not None else ''
                
                ws.cell(row=current_row, column=col_num, value=cell_value)
            
            # Now add detailed overtime row data (what you specifically want to see)
            form_id = form_data.get('id')
            if form_id:
                # Get overtime rows for this form
                c.execute('SELECT * FROM exception_form_rows WHERE form_id = ?', (form_id,))
                overtime_rows = c.fetchall()
                
                if overtime_rows:
                    # Get overtime row column names
                    overtime_columns = [desc[0] for desc in c.description]
                    
                    # Add overtime row headers if this is the first form with overtime data
                    if current_row == 2:
                        overtime_header_start = len(enhanced_columns) + 1
                        overtime_headers = [
                            'OT Row ID', 'Exception Code', 'Code Description', 'Line/Location', 'Run No.',
                            'Exception Time From (HH:MM)', 'Exception Time To (HH:MM)',
                            'Overtime Hours (HH:MM)', 'Bonus Hours (HH:MM)', 'Night Differential (HH:MM)',
                            'TA Job No.'
                        ]
                        
                        for col_num, header in enumerate(overtime_headers, overtime_header_start):
                            cell = ws.cell(row=1, column=col_num, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                    
                    # Add overtime row data
                    for overtime_row in overtime_rows:
                        overtime_data = dict(zip(overtime_columns, overtime_row))
                        
                        # Add overtime row data starting after the main form data
                        col_offset = len(enhanced_columns)
                        
                        # OT Row ID
                        ws.cell(row=current_row, column=col_offset + 1, value=overtime_data.get('id', ''))
                        
                        # Exception Code
                        ws.cell(row=current_row, column=col_offset + 2, value=overtime_data.get('code', ''))
                        
                        # Code Description (REASON for overtime)
                        ws.cell(row=current_row, column=col_offset + 3, value=overtime_data.get('code_description', ''))
                        
                        # Line/Location
                        ws.cell(row=current_row, column=col_offset + 4, value=overtime_data.get('line_location', ''))
                        
                        # Run No.
                        ws.cell(row=current_row, column=col_offset + 5, value=overtime_data.get('run_no', ''))
                        
                        # Exception Time From (HH:MM) - START TIME
                        time_from_hh = overtime_data.get('exception_time_from_hh', '')
                        time_from_mm = overtime_data.get('exception_time_from_mm', '')
                        time_from = f"{time_from_hh}:{time_from_mm}" if time_from_hh and time_from_mm else ''
                        ws.cell(row=current_row, column=col_offset + 6, value=time_from)
                        
                        # Exception Time To (HH:MM) - END TIME
                        time_to_hh = overtime_data.get('exception_time_to_hh', '')
                        time_to_mm = overtime_data.get('exception_time_to_mm', '')
                        time_to = f"{time_to_hh}:{time_to_mm}" if time_to_hh and time_to_mm else ''
                        ws.cell(row=current_row, column=col_offset + 7, value=time_to)
                        
                        # Overtime Hours (HH:MM) - DURATION
                        ot_hh = overtime_data.get('overtime_hh', '')
                        ot_mm = overtime_data.get('overtime_mm', '')
                        overtime_hours = f"{ot_hh}:{ot_mm}" if ot_hh and ot_mm else ''
                        ws.cell(row=current_row, column=col_offset + 8, value=overtime_hours)
                        
                        # Bonus Hours (HH:MM)
                        bonus_hh = overtime_data.get('bonus_hh', '')
                        bonus_mm = overtime_data.get('bonus_mm', '')
                        bonus_hours = f"{bonus_hh}:{bonus_mm}" if bonus_hh and bonus_mm else ''
                        ws.cell(row=current_row, column=col_offset + 9, value=bonus_hours)
                        
                        # Night Differential (HH:MM)
                        nite_hh = overtime_data.get('nite_diff_hh', '')
                        nite_mm = overtime_data.get('nite_diff_mm', '')
                        nite_diff = f"{nite_hh}:{nite_mm}" if nite_hh and nite_mm else ''
                        ws.cell(row=current_row, column=col_offset + 10, value=nite_diff)
                        
                        # TA Job No.
                        ws.cell(row=current_row, column=col_offset + 11, value=overtime_data.get('ta_job_no', ''))
                        
                        current_row += 1
                else:
                    current_row += 1
            else:
                current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generate filename based on filters
    filename_parts = ['exception_forms']
    if form_type:
        filename_parts.append(form_type)
    if extraction_mode:
        filename_parts.append(extraction_mode)
    filename = '_'.join(filename_parts) + '.xlsx'
    
    return Response(
        excel_buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename={filename}',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )

if __name__ == "__main__":
    init_audit_db()
    init_exception_form_db()
    app.run(port=8000, debug=True)

    
